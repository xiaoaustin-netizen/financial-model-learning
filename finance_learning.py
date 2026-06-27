import streamlit as st
import pandas as pd
import numpy as np
import re as _re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Financial Modeling School", layout="wide", page_icon="📈")

if "current_page" not in st.session_state:
    st.session_state.current_page = "overview"

# ── Practice companies ─────────────────────────────────────────────────────────
COMPANIES = [
    {"name":"NovaTech Industries",    "sector":"Technology",
     "revenue":"$500M",  "ebitda":"$120M", "ni":"$52.5M","price":"$45.00","shares":"20M",
     "rev_n":500, "ebitda_n":120, "ni_n":52.5, "price_n":45.00, "shares_n":20,
     "rev_growth":0.05, "wacc":0.0935, "terminal_g":0.025},
    {"name":"Meridian Capital Group", "sector":"Financial Services",
     "revenue":"$320M",  "ebitda":"$85M",  "ni":"$38M",  "price":"$28.00","shares":"15M",
     "rev_n":320, "ebitda_n":85,  "ni_n":38,   "price_n":28.00, "shares_n":15,
     "rev_growth":0.04, "wacc":0.0850, "terminal_g":0.020},
    {"name":"Apex Consumer Brands",   "sector":"Consumer Staples",
     "revenue":"$1.2B",  "ebitda":"$210M", "ni":"$95M",  "price":"$62.00","shares":"40M",
     "rev_n":1200,"ebitda_n":210, "ni_n":95,   "price_n":62.00, "shares_n":40,
     "rev_growth":0.03, "wacc":0.0750, "terminal_g":0.020},
    {"name":"Frontier Energy",        "sector":"Energy",
     "revenue":"$780M",  "ebitda":"$165M", "ni":"$72M",  "price":"$38.50","shares":"25M",
     "rev_n":780, "ebitda_n":165, "ni_n":72,   "price_n":38.50, "shares_n":25,
     "rev_growth":0.06, "wacc":0.1000, "terminal_g":0.020},
    {"name":"Cascade Logistics",      "sector":"Industrials",
     "revenue":"$440M",  "ebitda":"$95M",  "ni":"$41M",  "price":"$33.00","shares":"18M",
     "rev_n":440, "ebitda_n":95,  "ni_n":41,   "price_n":33.00, "shares_n":18,
     "rev_growth":0.05, "wacc":0.0875, "terminal_g":0.020},
]


def _co_inputs(co):
    """Compute per-page given-inputs from a company's numeric fields.
    IS, BS, and CFS figures are derived so they are internally consistent.
    Balance sheet always ties: Total Assets == Total L&E.
    """
    R, E, N, P, S = co["rev_n"], co["ebitda_n"], co["ni_n"], co["price_n"], co["shares_n"]
    TAX = 0.30

    # Income statement line items
    cogs     = round(R * 0.60, 1)
    sga      = round(R * 0.40 - E, 1)
    da       = round(R * 0.06, 1)
    ebit     = E - da
    interest = round(ebit - N / (1 - TAX), 1)
    capex    = round(R * 0.07, 1)
    divs     = round(N * 0.25, 1)

    # Net debt ~1.05× EBITDA, rounded to nearest $5M
    net_debt = round(E * 1.05 / 5) * 5

    # Balance sheet — built so Total Assets == Total L&E exactly
    cash   = round(R * 0.15, 1)
    ar     = round(R * 0.12, 1)
    inv    = round(R * 0.08, 1)
    ppe    = round(R * 0.40, 1)
    intang = round(R * 0.10, 1)
    ap     = round(R * 0.06, 1)
    st_debt = round(net_debt * 0.15 / 5) * 5        # ~15% of net debt, rounded to $5M
    lt_debt = net_debt + cash - st_debt              # ensures net_debt = LTD + STD − Cash
    total_assets = cash + ar + inv + ppe + intang
    total_liab   = ap + st_debt + lt_debt
    equity       = total_assets - total_liab
    com_stock    = round(equity * 0.50, 1)
    ret_earn     = round(equity - com_stock, 1)

    def fm(v):
        if abs(v) >= 1000:
            return f"${v/1000:.1f}B"
        return f"${int(v)}M" if v == int(v) else f"${v}M"

    ebitda_margin = E / R
    rev_growth = co.get("rev_growth", 0.05)
    wacc       = co.get("wacc", 0.0935)
    terminal_g = co.get("terminal_g", 0.025)

    # Budget vs. Actual — Q1 figures (~24% of annual)
    # Keep unit count constant (12,000 budget / 11,500 actual); price per unit scales with revenue.
    Q1_UNITS_B, Q1_UNITS_A = 12_000, 11_500
    q1_rev_b   = round(R * 0.24, 1)
    q1_rev_a   = round(q1_rev_b * 0.942, 1)
    q1_cogs_b  = round(cogs * 0.24, 1)
    q1_cogs_a  = round(q1_cogs_b * 0.976, 1)
    q1_sga_b   = round(sga * 0.25, 1)
    q1_sga_a   = round(q1_sga_b * 1.075, 1)
    q1_price_b = round(q1_rev_b * 1e6 / Q1_UNITS_B)
    q1_price_a = round(q1_rev_a * 1e6 / Q1_UNITS_A)

    # Merger target (always ~60% of acquirer size)
    tgt_price   = round(P * 0.60, 2)
    tgt_shares  = round(S * 0.40, 1)
    tgt_ni      = round(N * 0.35, 1)
    synergies   = round(N * 0.08, 1)
    intang_amort = round(N * 0.04, 1)

    return {
        "3stmt": [
            ("Revenue", fm(R)), ("COGS", fm(cogs)), ("SG&A", fm(sga)),
            ("D&A", fm(da)), ("Interest Exp.", fm(interest)), ("Tax Rate", "30%"),
            ("Beg. Cash", fm(cash)), ("Beg. AR", fm(ar)), ("Beg. Inventory", fm(inv)),
            ("Beg. PP&E", fm(ppe)), ("Beg. Intangibles", fm(intang)),
            ("Beg. AP", fm(ap)), ("Beg. ST Debt", fm(st_debt)), ("Beg. LT Debt", fm(lt_debt)),
            ("Beg. Com. Stock", fm(com_stock)), ("Beg. Ret. Earn.", fm(ret_earn)),
            ("Capex", fm(capex)), ("Dividends", fm(divs)),
        ],
        "dcf": [
            ("Base Revenue", fm(R)), ("Rev. Growth", f"{rev_growth*100:.1f}%"),
            ("EBITDA Margin", f"{ebitda_margin*100:.1f}%"), ("D&A % Rev.", "6.0%"),
            ("Capex % Rev.", "7.0%"), ("Tax Rate", "30%"),
            ("WACC", f"{wacc*100:.2f}%"), ("Terminal g", f"{terminal_g*100:.1f}%"),
            ("Net Debt", fm(net_debt)), ("Shares Out.", f"{int(S)}M"),
        ],
        "comps": [
            ("LTM EBITDA", fm(E)), ("Net Debt", fm(net_debt)), ("Shares Out.", f"{int(S)}M"),
        ],
        "precedent": [
            ("LTM EBITDA", fm(E)), ("Net Debt", fm(net_debt)),
            ("Shares Out.", f"{int(S)}M"), ("Current Price", f"${P:.2f}"),
        ],
        "lbo": [
            ("LTM EBITDA", fm(E)), ("Entry EV/EBITDA", "9.0×"),
            ("Debt / EV", "60%"), ("Interest Rate", "7.0%"),
            ("EBITDA Growth", "5% / yr"), ("D&A", fm(da)),
            ("Capex", fm(capex)), ("Tax Rate", "30%"), ("Exit EV/EBITDA", "9.0×"),
        ],
        "merger": [
            ("Target Price", f"${tgt_price:.2f}"), ("Acq. Premium", "25%"),
            ("Target Shares", f"{tgt_shares}M"), ("Target NI", fm(tgt_ni)),
            ("Acquirer NI", fm(N)), ("Acquirer Shares", f"{int(S)}M"),
            ("After-tax Synergies", fm(synergies)), ("Intangibles Amort.", f"${intang_amort}M/yr"),
        ],
        "budget": [
            ("Budg. Revenue", fm(q1_rev_b)), ("Act. Revenue", fm(q1_rev_a)),
            ("Budg. COGS", fm(q1_cogs_b)), ("Act. COGS", fm(q1_cogs_a)),
            ("Budg. SG&A", fm(q1_sga_b)), ("Act. SG&A", fm(q1_sga_a)),
            ("Budg. Units", f"{Q1_UNITS_B:,}"), ("Budg. Price/Unit", f"${q1_price_b:,}"),
            ("Act. Units", f"{Q1_UNITS_A:,}"), ("Act. Price/Unit", f"${q1_price_a:,}"),
        ],
    }
if "company_idx" not in st.session_state:
    st.session_state.company_idx = 0

def current_company():
    return COMPANIES[st.session_state.company_idx]

def nav_btn(label, key):
    active = st.session_state.current_page == key
    btn_type = "primary" if active else "secondary"
    if st.button(label, key=f"nav_{key}", use_container_width=True, type=btn_type):
        st.session_state.current_page = key
        st.rerun()

st.markdown("""<style>
/* ── Collapse Streamlit top bar to zero height (preserves sidebar toggle) ── */
header[data-testid="stHeader"] {
    height:0 !important; min-height:0 !important;
    padding:0 !important; margin:0 !important;
    overflow:visible !important;
}
/* Always keep the sidebar expand button reachable */
[data-testid="collapsedControl"] {
    display:flex !important; visibility:visible !important;
    opacity:1 !important; pointer-events:auto !important;
    z-index:999999 !important;
}

/* ── Global background & text ── */
.stApp { background:#F8FAFC; }
.stApp p, .stApp li, .stMarkdown p, .stMarkdown li { color:#374151; font-size:.93rem; line-height:1.72; }
.stApp h1 { color:#111827; font-size:1.65rem; font-weight:700; letter-spacing:-.01em; }
.stApp h2 { color:#111827; font-size:1.15rem; font-weight:700; margin-top:1.8rem; }
.stApp h3 { color:#1F2937; font-size:1rem; font-weight:600; }
.stApp strong, .stMarkdown strong { color:#111827; }
.block-container { padding-top:0 !important; padding-bottom:2.5rem; max-width:1100px; }

/* ── Page header ── */
.page-hdr-title { font-size:1.08rem; font-weight:800; color:#111827; letter-spacing:-.01em; line-height:1.25; }
.page-hdr-sub   { font-size:.72rem; color:#9CA3AF; margin-top:2px; }
.page-hdr-co    {
    display:inline-flex; align-items:center; gap:6px;
    padding:5px 12px; background:#EFF6FF;
    border:1px solid #BFDBFE; border-radius:5px;
}
.page-hdr-co-name   { font-size:.78rem; font-weight:700; color:#1D4ED8; }
.page-hdr-co-sector { font-size:.65rem; color:#64748B; }
.page-hdr-inputs-lbl {
    font-size:.57rem; font-weight:700; letter-spacing:.14em;
    text-transform:uppercase; color:#94A3B8; margin-bottom:5px;
}

/* ── Input chips ── */
.input-chip {
    display:inline-flex; align-items:baseline; gap:4px;
    font-size:.71rem; color:#111827; font-weight:600;
    background:#F8FAFC; padding:2px 8px;
    border:1px solid #E2E8F0; border-radius:3px;
    margin:2px 3px 2px 0; white-space:nowrap;
}
.input-chip-lbl {
    font-size:.57rem; color:#94A3B8; font-weight:500;
    letter-spacing:.03em; text-transform:uppercase;
}

/* ── Section labels ── */
.sec {
    font-size:.58rem; font-weight:700; letter-spacing:.16em;
    text-transform:uppercase; color:#9CA3AF;
    border-bottom:1px solid #E5E7EB; padding-bottom:6px; margin-bottom:14px;
}

/* ── Cards ── */
.card {
    border-radius:6px; padding:16px 22px; margin-bottom:14px;
    border-left:3px solid; box-shadow:0 1px 2px rgba(0,0,0,.05);
    color:#1F2937; font-size:.88rem; line-height:1.65;
}
.card-concept { background:#EFF6FF; border-color:#2563EB; }
.card-formula { background:#F5F3FF; border-color:#7C3AED; }
.card-warning { background:#FFFBEB; border-color:#D97706; }
.card-pro     { background:#F9FAFB; border-color:#D1D5DB; }
.card b, .card strong { color:#111827; }
.card-title {
    font-size:.58rem; font-weight:700; letter-spacing:.14em;
    text-transform:uppercase; margin-bottom:8px;
}
.ct-concept { color:#1D4ED8; }
.ct-formula { color:#6D28D9; }
.ct-warning { color:#B45309; }
.ct-pro     { color:#6B7280; }
.formula-text {
    font-family:'SF Mono','Fira Code','Cascadia Code',monospace;
    font-size:.92rem; color:#1F2937; font-weight:600;
    background:#FFFFFF; border:1px solid #E5E7EB;
    padding:8px 12px; display:block; margin-top:4px;
}

/* ── Overview landing cards ── */
.lp-card {
    background:#FFFFFF; border:1px solid #E5E7EB;
    border-radius:6px; padding:18px 22px; margin-bottom:8px; border-top:3px solid;
    box-shadow:0 1px 3px rgba(0,0,0,.04);
}
.lp-title { font-size:.93rem; font-weight:700; color:#111827; }
.lp-desc  { font-size:.82rem; color:#4B5563; margin-top:7px; line-height:1.6; }

/* ── Guide panel for Build tabs ── */
.guide-panel {
    background:#F8FAFC; border:1px solid #E5E7EB;
    border-left:3px solid #2563EB;
    padding:12px 14px; margin-bottom:8px;
}
.guide-line {
    font-size:.79rem; color:#374151; line-height:1.85;
    padding:3px 0; border-bottom:1px solid #F3F4F6;
}
.guide-line:last-child { border-bottom:none; }

/* ── Reveal-answer expanders (compact, subtle) ── */
.guide-line + [data-testid="stExpander"],
[data-testid="stExpander"]:has(div[style*="#FEFCE8"]) {
    margin-top:0 !important; margin-bottom:2px !important;
}
[data-testid="stExpander"] > details > summary {
    font-size:.65rem !important;
    color:#9CA3AF !important;
    padding:2px 6px !important;
    min-height:0 !important;
    background:transparent !important;
    border:1px solid #E5E7EB !important;
    border-radius:4px !important;
}
[data-testid="stExpander"] > details > summary:hover {
    color:#6B7280 !important;
    border-color:#D1D5DB !important;
}
[data-testid="stExpander"] > details > summary svg {
    width:10px !important; height:10px !important;
}
[data-testid="stExpander"] > details[open] > summary {
    color:#CA8A04 !important;
    border-color:#FDE68A !important;
    background:#FFFBEB !important;
}

/* ── Sidebar — grey ladder ── */
section[data-testid="stSidebar"],
[data-testid="stSidebar"] {
    background:#F9FAFB !important;
    border-right:1px solid #E5E7EB !important;
}
section[data-testid="stSidebar"] > div { padding-top:0 !important; }

.nav-wordmark {
    background:#F3F4F6; border-bottom:1px solid #E5E7EB;
    padding:15px 16px; margin-bottom:4px;
    font-size:.68rem; font-weight:800; letter-spacing:.22em;
    text-transform:uppercase; color:#374151;
}
.nav-group-hdr {
    font-size:.53rem; font-weight:700; letter-spacing:.2em;
    text-transform:uppercase; color:#9CA3AF;
    padding:14px 16px 3px; margin:0; border:none;
}

/* ── Sidebar buttons: nuclear override ── */
[data-testid="stSidebar"] .stButton > button,
[data-testid="stSidebar"] .stButton > button:focus,
[data-testid="stSidebar"] .stButton > button:active,
[data-testid="stSidebar"] .stButton > button:focus-visible,
[data-testid="stSidebar"] .stButton > button:focus:not(:active) {
    width:100% !important; text-align:left !important;
    border-radius:0 !important;
    border-top:none !important; border-right:none !important;
    border-bottom:1px solid #E5E7EB !important;
    border-left:3px solid #C9CDD4 !important;
    outline:none !important; box-shadow:none !important;
    font-size:.82rem !important; font-weight:500 !important;
    padding:9px 12px 9px 14px !important; margin:0 !important;
    background-color:#EAECEF !important; color:#1E3A5F !important;
}
[data-testid="stSidebar"] .stButton > button *,
[data-testid="stSidebar"] .stButton > button p,
[data-testid="stSidebar"] .stButton > button span,
[data-testid="stSidebar"] .stButton > button div { color:#1E3A5F !important; }
[data-testid="stSidebar"] .stButton > button:hover {
    background-color:#D9DCE0 !important; border-left-color:#6B7280 !important; color:#111827 !important;
}
[data-testid="stSidebar"] .stButton > button:hover * { color:#111827 !important; }
[data-testid="stSidebar"] .stButton > button[kind="primary"],
[data-testid="stSidebar"] .stButton > button[kind="primary"]:focus,
[data-testid="stSidebar"] .stButton > button[kind="primary"]:active {
    background-color:#D1D5DB !important; border-left-color:#1E3A5F !important;
    color:#111827 !important; font-weight:700 !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] * {
    color:#111827 !important; font-weight:700 !important;
}

/* ── Hint rows ── */
.hint-row { font-size:.8rem; color:#4B5563; padding:6px 0; border-bottom:1px solid #F3F4F6; }
.hint-num { display:inline-block; width:54px; font-family:monospace; color:#9CA3AF; font-size:.71rem; }

/* ── Tabs ── */
button[data-baseweb="tab"] { font-size:.82rem; font-weight:600; color:#4B5563; }

/* ── Fix dollar-sign LaTeX rendering in learn tabs ── */
.stMarkdown .katex,
.stMarkdown .katex-html,
.stMarkdown .katex-error {
    background: transparent !important;
    color: inherit !important;
    font-size: inherit !important;
    font-family: inherit !important;
    border: none !important;
    padding: 0 !important;
}

/* ── Reroll (↻) company button — light/subtle style ── */
div[data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"],
div[data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"]:focus,
div[data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"]:active {
    background: #F9FAFB !important;
    border: 1px solid #D1D5DB !important;
    color: #374151 !important;
    border-radius: 6px !important;
    font-size: .95rem !important;
    box-shadow: none !important;
}
div[data-testid="stHorizontalBlock"] .stButton > button[kind="secondary"]:hover {
    background: #F3F4F6 !important;
    border-color: #9CA3AF !important;
    color: #111827 !important;
}
</style>""", unsafe_allow_html=True)

# ── Card helpers ───────────────────────────────────────────────────────────────
def concept(title, body):
    st.markdown(f'<div class="card card-concept"><div class="card-title ct-concept">Concept — {title}</div>{body}</div>', unsafe_allow_html=True)

def formula_card(expr, note=""):
    note_html = f'<div style="font-size:.78rem;color:#6D28D9;margin-top:8px;">{note}</div>' if note else ""
    st.markdown(f'<div class="card card-formula"><div class="card-title ct-formula">Formula</div><div class="formula-text">{expr}</div>{note_html}</div>', unsafe_allow_html=True)

def warning(body):
    st.markdown(f'<div class="card card-warning"><div class="card-title ct-warning">Common Mistake</div>{body}</div>', unsafe_allow_html=True)

def pro_tip(body):
    st.markdown(f'<div class="card card-pro"><div class="card-title ct-pro">Analyst Note</div>{body}</div>', unsafe_allow_html=True)

def given_inputs(_items, _note=""):
    pass

EXCEL_LINKS = {
    "3stmt":     "",  # Three Statement Model — paste spreadsheet URL here
    "dcf":       "",  # DCF Model — paste spreadsheet URL here
    "comps":     "",  # Comps — paste spreadsheet URL here
    "precedent": "",  # Precedent Transactions — paste spreadsheet URL here
    "lbo":       "",  # LBO Model — paste spreadsheet URL here
    "merger":    "",  # Merger Model — paste spreadsheet URL here
    "budget":    "",  # Budget vs. Actual — paste spreadsheet URL here
}


def _excel_link_tab(model_key, model_name):
    """Render a card linking to the external Excel template for this model."""
    url = EXCEL_LINKS.get(model_key, "")
    st.markdown('<div style="height:16px;"></div>', unsafe_allow_html=True)
    if url:
        st.markdown(
            f'<div style="background:#EFF6FF;border:1px solid #BFDBFE;'
            f'border-left:4px solid #2563EB;border-radius:6px;'
            f'padding:28px 36px;text-align:center;">'
            f'<div style="font-size:.65rem;font-weight:700;letter-spacing:.18em;'
            f'text-transform:uppercase;color:#1D4ED8;margin-bottom:10px;">Excel Template</div>'
            f'<div style="font-size:.97rem;color:#1F2937;font-weight:600;margin-bottom:20px;">'
            f'Build the {model_name} in Excel</div>'
            f'<a href="{url}" target="_blank" rel="noopener noreferrer" '
            f'style="display:inline-block;background:#2563EB;color:#FFFFFF;'
            f'font-size:.85rem;font-weight:600;padding:11px 32px;'
            f'border-radius:5px;text-decoration:none;">'
            f'Open Excel Template →</a>'
            f'</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="background:#F9FAFB;border:1.5px dashed #D1D5DB;'
            f'border-radius:6px;padding:28px 36px;text-align:center;">'
            f'<div style="font-size:.88rem;font-weight:600;color:#374151;margin-bottom:8px;">'
            f'Excel Template — Link Not Configured</div>'
            f'<div style="font-size:.80rem;color:#9CA3AF;line-height:1.7;">'
            f'Open <code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;">'
            f'finance_learning.py</code> and set the URL for '
            f'<code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;">'
            f'{model_key}</code> in the '
            f'<code style="background:#F3F4F6;padding:1px 5px;border-radius:3px;">EXCEL_LINKS</code>'
            f' dict near the top of the file.</div>'
            f'</div>',
            unsafe_allow_html=True)


# ── Formula engine ─────────────────────────────────────────────────────────────

def _eval_formula(raw, cell_vals, cross=None):
    """Evaluate one cell. cross = {SheetName: {ref: val}} for cross-section refs."""
    if cross is None:
        cross = {}
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return ""
    if not s.startswith("="):
        clean = s.replace("$", "").replace(",", "")
        try:
            return float(clean)
        except:
            return s

    expr = s[1:]

    # Strip $ from refs: $B$2 → B2
    expr = _re.sub(r'\$([A-Z])\$(\d+)', r'\1\2', expr)
    expr = _re.sub(r'\$([A-Z])(\d+)', r'\1\2', expr)
    expr = _re.sub(r'([A-Z])\$(\d+)', r'\1\2', expr)

    # Cross-sheet: IS!B12 → value from cross["IS"]["B12"]
    def _xr(m):
        sheet = m.group(1).upper()
        ref = m.group(2).upper()
        v = cross.get(sheet, {}).get(ref, 0)
        return str(float(v)) if isinstance(v, (int, float)) else "0"
    expr = _re.sub(r'([A-Za-z_]\w*)!([A-Z]\d+)', _xr, expr)

    # SUM(A1:A5)
    def _sum(m):
        col, r1, r2 = m.group(1).upper(), int(m.group(2)), int(m.group(4))
        total = sum(float(cell_vals.get(f"{col}{r}", 0) or 0)
                    for r in range(r1, r2+1)
                    if isinstance(cell_vals.get(f"{col}{r}", 0), (int, float)))
        return str(total)
    expr = _re.sub(r'SUM\(([A-Z])(\d+):([A-Z])(\d+)\)', _sum, expr, flags=_re.I)

    # Aggregate functions
    for fn_name, fn_np in [("MEDIAN", np.median), ("AVERAGE", np.mean), ("MAX", np.max), ("MIN", np.min)]:
        def _agg(m, fn=fn_np):
            col, r1, r2 = m.group(1).upper(), int(m.group(2)), int(m.group(4))
            vs = [float(cell_vals.get(f"{col}{r}", 0) or 0)
                  for r in range(r1, r2+1)
                  if isinstance(cell_vals.get(f"{col}{r}", 0), (int, float))]
            return str(round(float(fn(vs)), 8)) if vs else "0"
        expr = _re.sub(fr'{fn_name}\(([A-Z])(\d+):([A-Z])(\d+)\)', _agg, expr, flags=_re.I)

    # PERCENTILE(A1:A5, 0.25)
    def _pct(m):
        col, r1, r2, p = m.group(1).upper(), int(m.group(2)), int(m.group(4)), float(m.group(5))
        vs = [float(cell_vals.get(f"{col}{r}", 0) or 0)
              for r in range(r1, r2+1)
              if isinstance(cell_vals.get(f"{col}{r}", 0), (int, float))]
        return str(round(np.percentile(vs, p * 100), 8)) if vs else "0"
    expr = _re.sub(r'PERCENTILE\(([A-Z])(\d+):([A-Z])(\d+),\s*([0-9.]+)\)', _pct, expr, flags=_re.I)

    # IF(cond, true, false) — simple, no nesting
    def _if(m):
        cond_s, tv, fv = m.group(1).strip(), m.group(2).strip().strip('"\''), m.group(3).strip().strip('"\'')
        ce = _re.sub(r'[A-Z]\d+', lambda mc: str(float(cell_vals.get(mc.group(0).upper(), 0) or 0)), cond_s)
        try:
            return f'"{tv}"' if bool(eval(ce, {"__builtins__": {}}, {})) else f'"{fv}"'
        except:
            return f'"{fv}"'
    expr = _re.sub(r'IF\(([^,]+),([^,]+),([^)]+)\)', _if, expr, flags=_re.I)

    # IRR(B0:F0) — Newton-Raphson
    def _irr(m):
        col1, row, col2 = m.group(1).upper(), int(m.group(2)), m.group(3).upper()
        cols = [chr(c) for c in range(ord(col1), ord(col2)+1)]
        flows = [float(cell_vals.get(f"{c}{row}", 0) or 0) for c in cols]
        rate = 0.1
        for _ in range(200):
            try:
                npv = sum(f / (1+rate)**i for i, f in enumerate(flows))
                dnpv = -sum(i*f / (1+rate)**(i+1) for i, f in enumerate(flows) if i > 0)
                if abs(dnpv) < 1e-12:
                    break
                rate = max(rate - npv/dnpv, -0.9999)
            except:
                return "#IRR"
        return str(round(rate, 6))
    expr = _re.sub(r'IRR\(([A-Z])(\d+):([A-Z])(\d+)\)', _irr, expr, flags=_re.I)

    # Cell references
    def _ref(m):
        v = cell_vals.get(m.group(0).upper(), 0)
        return str(float(v)) if isinstance(v, (int, float)) else "0"
    expr = _re.sub(r'[A-Z]\d+', _ref, expr)

    expr = expr.replace("^", "**")
    try:
        result = eval(expr, {"__builtins__": {}}, {})
        if isinstance(result, str):
            return result.strip('"\'')
        return round(float(result), 6)
    except:
        return "#ERR"


def _compute_grid(b_vals, cross=None, seed=None):
    """Evaluate all B-column cells; seed = extra {ref: val} to pre-populate."""
    cell_vals = dict(seed or {})
    bs = list(b_vals)
    # Literals first
    for i, b in enumerate(bs):
        s = str(b).strip()
        if s and s.lower() != "nan" and not s.startswith("="):
            try:
                cell_vals[f"B{i+1}"] = float(s.replace("$","").replace(",",""))
            except:
                cell_vals[f"B{i+1}"] = s
    # Multi-pass formulas
    for _ in range(8):
        for i, b in enumerate(bs):
            if str(b).strip().startswith("="):
                cell_vals[f"B{i+1}"] = _eval_formula(b, cell_vals, cross)
    return cell_vals


def _compute_multicol(col_data, cross=None, seed=None):
    """
    col_data: dict {col_letter: [val_list]} e.g. {"B": [...], "C": [...]}
    Returns merged cell_vals dict with B1, B2, C1, C2, etc.
    """
    cell_vals = dict(seed or {})
    # Literals
    for col, vals in col_data.items():
        for i, v in enumerate(vals):
            s = str(v).strip()
            if s and s.lower() != "nan" and not s.startswith("="):
                try:
                    cell_vals[f"{col.upper()}{i+1}"] = float(s.replace("$","").replace(",",""))
                except:
                    cell_vals[f"{col.upper()}{i+1}"] = s
    # Multi-pass formulas
    for _ in range(8):
        for col, vals in col_data.items():
            for i, v in enumerate(vals):
                if str(v).strip().startswith("="):
                    cell_vals[f"{col.upper()}{i+1}"] = _eval_formula(v, cell_vals, cross)
    return cell_vals


def _fmt(v):
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e10:
            return f"{int(v):,}"
        if abs(v) < 1:
            return f"{v:.4f}".rstrip('0').rstrip('.')
        return f"{v:,.2f}"
    return str(v) if (v is not None and str(v) not in ("", "nan")) else ""


def _get_all_cells():
    if "_all_cells" not in st.session_state:
        st.session_state._all_cells = {}
    return st.session_state._all_cells



def formula_grid_section(section_key, title, guidance_lines, n_rows=12,
                          cross=None, seed=None, section_name=None, reveal_lines=None):
    """
    Two-column guided grid: left = guidance, right = Excel-like formula grid.
    Formula bar: name-box (cell selector) + fx input + apply button.
    Function dropdown inserts template into active cell.
    Cell-ref chips insert reference into active cell formula.
    reveal_lines: optional parallel list; per-row "show answer" button reveals this text.
    """
    sk = f"fgs_{section_key}"
    if sk not in st.session_state:
        st.session_state[sk] = {"a": [""] * n_rows, "b": [""] * n_rows, "ver": 0, "active": 1}
    s = st.session_state[sk]
    while len(s["a"]) < n_rows: s["a"].append("")
    while len(s["b"]) < n_rows: s["b"].append("")
    if "active" not in s: s["active"] = 1

    cell_vals = _compute_grid(s["b"], cross=cross, seed=seed)
    if section_name:
        _get_all_cells()[section_name] = cell_vals

    st.markdown(f'<div class="sec">{title}</div>', unsafe_allow_html=True)

    hints_on = st.session_state.get("hints_on", True)
    active = s["active"]

    if hints_on:
        col_g, col_grid = st.columns([5, 8])
        with col_g:
            for idx, line in enumerate(guidance_lines):
                rn = idx + 1
                is_act = (rn == active)
                bg = "background:#EFF6FF;" if is_act else ""
                bl = "border-left:3px solid #2563EB;" if is_act else "border-left:3px solid #E5E7EB;"
                bc = "#1D4ED8" if is_act else "#94A3B8"
                badge = (f'<span style="font-family:monospace;font-weight:700;font-size:.59rem;'
                         f'color:{bc};min-width:26px;display:inline-block;flex-shrink:0;'
                         f'padding-top:1px;">B{rn}</span>')
                st.markdown(
                    f'<div class="guide-line" style="{bg}{bl}padding:4px 6px;'
                    f'display:flex;align-items:flex-start;gap:5px;border-bottom:1px solid #F3F4F6;">'
                    f'{badge}<span style="font-size:.74rem;">{line}</span></div>',
                    unsafe_allow_html=True)
                if reveal_lines and idx < len(reveal_lines):
                    with st.expander("show answer"):
                        st.markdown(
                            f'<div style="background:#FEFCE8;border-left:3px solid #CA8A04;'
                            f'border-radius:0 3px 3px 0;padding:5px 8px;font-size:.73rem;'
                            f'color:#713F12;">{reveal_lines[idx]}</div>',
                            unsafe_allow_html=True)
    else:
        col_grid = st.container()

    with col_grid:
        ver = s["ver"]
        cur_b = s["b"][active - 1] if active <= len(s["b"]) else ""
        fbar_key = f"xlfbar_{section_key}_v{ver}_r{active}"

        # ── Excel-style formula bar ──────────────────────────
        # Name-box (cell selector) | formula input | ✓ commit | ↺ reset
        cb1, cb2, cb3, cb4 = st.columns([1.1, 6.8, 0.55, 0.55])
        with cb1:
            sel = st.selectbox("cell", [f"B{i+1}" for i in range(n_rows)],
                               index=active - 1, label_visibility="collapsed",
                               key=f"xlcell_{section_key}_v{ver}")
            new_active = int(sel[1:])
            if new_active != active:
                s["active"] = new_active
                st.rerun()
        with cb2:
            formula_input = st.text_input(
                "fx", value=cur_b, key=fbar_key,
                placeholder="type a value  or  =formula …",
                label_visibility="collapsed")
        with cb3:
            if st.button("✓", key=f"xlapply_{section_key}_v{ver}",
                         use_container_width=True, help="Commit formula to cell (Enter)"):
                s["b"][active - 1] = formula_input
                s["ver"] += 1; st.rerun()
        with cb4:
            if st.button("↺", key=f"xlclr_{section_key}_v{ver}",
                         use_container_width=True, help="Clear and reset entire sheet"):
                s["a"] = [""] * n_rows; s["b"] = [""] * n_rows
                s["ver"] += 1; s["active"] = 1; st.rerun()

        # ── Function shortcut buttons ────────────────────────
        # Each button appends the function call to whatever is in the formula input,
        # or inserts a full template when the cell is empty.
        FN_BTNS = [
            ("SUM",     "SUM(",        "=SUM(B1:B5)"),
            ("IF",      "IF(",         "=IF(B1>0,1,0)"),
            ("AVG",     "AVERAGE(",    "=AVERAGE(B1:B5)"),
            ("MAX",     "MAX(",        "=MAX(B1:B5)"),
            ("MIN",     "MIN(",        "=MIN(B1:B5)"),
            ("MEDIAN",  "MEDIAN(",     "=MEDIAN(B1:B5)"),
            ("%ILE",    "PERCENTILE(", "=PERCENTILE(B1:B5,0.25)"),
            ("IRR",     "IRR(",        "=IRR(B1:F1)"),
        ]
        st.markdown(
            '<div style="font-size:.58rem;font-weight:700;letter-spacing:.1em;'
            'text-transform:uppercase;color:#6B7280;margin:5px 0 3px;">ƒ  Functions</div>',
            unsafe_allow_html=True)
        fn_cols = st.columns(len(FN_BTNS))
        for col_w, (lbl, fn_call, tpl) in zip(fn_cols, FN_BTNS):
            with col_w:
                if st.button(lbl, key=f"fn_{lbl}_{section_key}_v{ver}",
                             use_container_width=True, help=f"Insert  {tpl}"):
                    prog = str(st.session_state.get(fbar_key, cur_b) or "")
                    s["b"][active - 1] = (prog + fn_call) if prog.startswith("=") else tpl
                    s["ver"] += 1; st.rerun()

        # ── Click-to-reference chips ─────────────────────────
        # Shows every populated cell; clicking appends its reference to the formula.
        filled = [(f"B{i+1}", cell_vals.get(f"B{i+1}", ""))
                  for i in range(n_rows) if s["b"][i] and str(s["b"][i]).strip()]
        if filled:
            st.markdown(
                '<div style="font-size:.58rem;font-weight:700;letter-spacing:.1em;'
                'text-transform:uppercase;color:#6B7280;margin:5px 0 3px;">↩  Click cell to add reference</div>',
                unsafe_allow_html=True)
            ref_cols = st.columns(min(len(filled), 9))
            for i, (ref, val) in enumerate(filled[:9]):
                with ref_cols[i]:
                    tip = f"{ref} = {_fmt(val)}" if isinstance(val, (int, float)) else ref
                    if st.button(ref, key=f"xlref_{ref}_{section_key}_v{ver}",
                                 use_container_width=True, help=tip):
                        prog = str(st.session_state.get(fbar_key, cur_b) or "")
                        s["b"][active - 1] = (prog + ref) if prog.startswith("=") else f"={ref}"
                        s["ver"] += 1; st.rerun()

        # ── Grid — cells show computed results ───────────────
        def _disp(i):
            raw = s["b"][i]
            if not raw or str(raw).strip() == "":
                return ""
            if str(raw).strip().startswith("="):
                v = cell_vals.get(f"B{i+1}", "")
                return _fmt(v) if v not in ("", None) else raw
            return str(raw)

        display_b = [_disp(i) for i in range(n_rows)]

        rows = []
        for i in range(n_rows):
            a = s["a"][i]
            row_num = f"▶{i+1}" if (i + 1 == active) else str(i + 1)
            rows.append({" ": row_num, "A  ·  Label": a or "", "B  ·  Value": display_b[i]})

        df = pd.DataFrame(rows)
        ek = f"de_{section_key}_v{ver}"
        edited = st.data_editor(df, key=ek, hide_index=True, use_container_width=True,
            num_rows="fixed",
            column_config={
                " ":           st.column_config.TextColumn(" ", disabled=True, width=32),
                "A  ·  Label": st.column_config.TextColumn("A  ·  Label", width=148),
                "B  ·  Value": st.column_config.TextColumn("B  ·  Value", width=260),
            })

        na = [("" if str(x) in ("nan","None","") else str(x)) for x in edited["A  ·  Label"]]
        nb_edited = [("" if str(x) in ("nan","None","") else str(x)) for x in edited["B  ·  Value"]]

        new_b = list(s["b"])
        any_b_changed = False
        for i in range(n_rows):
            if nb_edited[i] != display_b[i]:
                new_b[i] = nb_edited[i]
                any_b_changed = True

        if na != s["a"] or any_b_changed:
            s["a"] = na
            s["b"] = new_b
            s["ver"] += 1
            st.rerun()


def formula_grid_multicol(section_key, title, guidance_lines, n_rows, col_labels,
                           cross=None, seed=None, section_name=None, reveal_lines=None):
    """
    Multi-column formula grid (e.g. Year 1–5 projections).
    col_labels: list of column header strings (one per value column, max 6).
    """
    cols = list("BCDEFG")[:len(col_labels)]
    sk = f"fgsm_{section_key}"
    if sk not in st.session_state:
        st.session_state[sk] = {
            "a": [""] * n_rows,
            "cols": {c: [""] * n_rows for c in cols},
            "ver": 0
        }
    s = st.session_state[sk]
    while len(s["a"]) < n_rows: s["a"].append("")
    for c in cols:
        if c not in s["cols"]: s["cols"][c] = [""] * n_rows
        while len(s["cols"][c]) < n_rows: s["cols"][c].append("")

    cell_vals = _compute_multicol(s["cols"], cross=cross, seed=seed)
    if section_name:
        _get_all_cells()[section_name] = cell_vals

    st.markdown(f'<div class="sec">{title}</div>', unsafe_allow_html=True)

    hints_on_mc = st.session_state.get("hints_on", True)

    if hints_on_mc:
        col_g, col_grid = st.columns([5, 8])
        with col_g:
            for idx, line in enumerate(guidance_lines):
                rn = idx + 1
                bg = "background:#EFF6FF;" if rn == 1 else ""
                bl = "border-left:3px solid #2563EB;" if rn == 1 else "border-left:3px solid #E5E7EB;"
                bc = "#1D4ED8" if rn == 1 else "#94A3B8"
                badge = (f'<span style="font-family:monospace;font-weight:700;font-size:.59rem;'
                         f'color:{bc};min-width:26px;display:inline-block;flex-shrink:0;'
                         f'padding-top:1px;">R{rn}</span>')
                st.markdown(
                    f'<div class="guide-line" style="{bg}{bl}padding:4px 6px;'
                    f'display:flex;align-items:flex-start;gap:5px;border-bottom:1px solid #F3F4F6;">'
                    f'{badge}<span style="font-size:.74rem;">{line}</span></div>',
                    unsafe_allow_html=True)
                if reveal_lines and idx < len(reveal_lines):
                    with st.expander("show answer"):
                        st.markdown(
                            f'<div style="background:#FEFCE8;border-left:3px solid #CA8A04;'
                            f'border-radius:0 3px 3px 0;padding:5px 8px;font-size:.73rem;'
                            f'color:#713F12;">{reveal_lines[idx]}</div>',
                            unsafe_allow_html=True)
    else:
        col_grid = st.container()

    with col_grid:
        # Cells show computed results inline; type =formula or value directly
        def _disp_mc(c, i):
            raw = s["cols"][c][i]
            if not raw or str(raw).strip() == "":
                return ""
            if str(raw).strip().startswith("="):
                v = cell_vals.get(f"{c.upper()}{i+1}", "")
                return _fmt(v) if v not in ("", None) else raw
            return str(raw)

        display_cols = {c: [_disp_mc(c, i) for i in range(n_rows)] for c in cols}

        rows = []
        for i in range(n_rows):
            row = {"#": str(i+1), "Label (A)": s["a"][i] or ""}
            for c, lbl in zip(cols, col_labels):
                row[lbl] = display_cols[c][i]
            rows.append(row)
        df = pd.DataFrame(rows)

        col_cfg = {
            "#": st.column_config.TextColumn("#", disabled=True, width=32),
            "Label (A)": st.column_config.TextColumn("Label (A)", width=130),
        }
        for lbl in col_labels:
            col_cfg[lbl] = st.column_config.TextColumn(lbl, width=130)

        ek = f"dem_{section_key}_v{s['ver']}"
        edited = st.data_editor(df, key=ek, hide_index=True, use_container_width=True,
                                num_rows="fixed", column_config=col_cfg)

        na = [("" if str(x) in ("nan","None","") else str(x)) for x in edited["Label (A)"]]

        # Detect changes: only update cells the user actually edited
        any_changed = (na != s["a"])
        new_cols = {c: list(s["cols"][c]) for c in cols}
        for c, lbl in zip(cols, col_labels):
            nc = [("" if str(x) in ("nan","None","") else str(x)) for x in edited[lbl]]
            for i in range(n_rows):
                if nc[i] != display_cols[c][i]:
                    new_cols[c][i] = nc[i]
                    any_changed = True

        if any_changed:
            s["a"] = na
            for c in cols:
                s["cols"][c] = new_cols[c]
            s["ver"] += 1
            st.rerun()


# ── Excel download helpers ──────────────────────────────────────────────────────

_BLUE_FILL  = PatternFill("solid", fgColor="BFDBFE")
_YELLOW_FILL= PatternFill("solid", fgColor="FFFBEB")
_HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
_SUB_FILL   = PatternFill("solid", fgColor="DBEAFE")
_BLUE_FONT  = Font(color="1D4ED8")
_GREEN_FONT = Font(color="065F46")
_BOLD_FONT  = Font(bold=True)
_BOLD_BL    = Font(bold=True, color="1E3A5F")
_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_RIGHT_ALIGN= Alignment(horizontal="right")
_CENTER_ALG = Alignment(horizontal="center")


def _xl_hdr_row(ws, row, *labels):
    for i, lbl in enumerate(labels, 1):
        col = chr(ord("A") + i - 1)
        c = ws[f"{col}{row}"]
        c.value = lbl; c.font = _WHITE_BOLD
        c.fill = _HDR_FILL; c.alignment = _CENTER_ALG


def _xl_given(ws, row, label, value, lc="A", vc="B"):
    ws[f"{lc}{row}"].value = label
    ws[f"{vc}{row}"].value = value
    ws[f"{vc}{row}"].fill = _BLUE_FILL
    ws[f"{vc}{row}"].font = _BLUE_FONT
    ws[f"{vc}{row}"].alignment = _RIGHT_ALIGN


def _xl_fml(ws, row, label, formula, lc="A", vc="B"):
    ws[f"{lc}{row}"].value = label
    ws[f"{vc}{row}"].value = formula
    ws[f"{vc}{row}"].font = _GREEN_FONT
    ws[f"{vc}{row}"].alignment = _RIGHT_ALIGN


def _xl_sub(ws, row, label, formula, lc="A", vc="B"):
    ws[f"{lc}{row}"].value = label
    ws[f"{lc}{row}"].font = _BOLD_BL
    ws[f"{vc}{row}"].value = formula
    ws[f"{vc}{row}"].font = _BOLD_BL
    ws[f"{vc}{row}"].fill = _SUB_FILL
    ws[f"{vc}{row}"].alignment = _RIGHT_ALIGN


def _xl_prac(ws, row, label, value=None, lc="A", vc="B"):
    ws[f"{lc}{row}"].value = label
    if value is not None:
        ws[f"{vc}{row}"].value = value
        ws[f"{vc}{row}"].fill = _BLUE_FILL
        ws[f"{vc}{row}"].font = _BLUE_FONT
        ws[f"{vc}{row}"].alignment = _RIGHT_ALIGN
    else:
        ws[f"{vc}{row}"].fill = _YELLOW_FILL


def _xl_sec(ws, row, label):
    ws[f"A{row}"].value = label
    ws[f"A{row}"].font = _BOLD_BL
    ws[f"A{row}"].fill = _SUB_FILL


def _cw(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def _make_3stmt_excel(wb, co):
    R = co["rev_n"]; E = co["ebitda_n"]; N = co["ni_n"]
    TAX = 0.30
    cogs = round(R * 0.60, 1);  sga = round(R * 0.40 - E, 1)
    da   = round(R * 0.06, 1);  ebit = E - da
    interest = round(ebit - N / (1 - TAX), 1)
    capex = round(R * 0.07, 1); divs = round(N * 0.25, 1)
    net_debt = round(E * 1.05 / 5) * 5
    cash = round(R*0.15,1); ar = round(R*0.12,1); inv = round(R*0.08,1)
    ppe  = round(R*0.40,1); intang = round(R*0.10,1)
    ap = round(R*0.06,1); st_debt = round(net_debt*0.15/5)*5
    lt_debt = net_debt + cash - st_debt
    tot_a = cash+ar+inv+ppe+intang; tot_l = ap+st_debt+lt_debt
    eq = tot_a - tot_l; com_stock = round(eq*0.50,1); ret_earn = round(eq-com_stock,1)
    nm = co["name"]

    # IS — Complete
    wc = wb.create_sheet("IS — Complete"); _cw(wc, [("A",28),("B",16)])
    _xl_hdr_row(wc, 1, f"Income Statement — {nm}", "($M)")
    _xl_given(wc,2,"Revenue",R); _xl_given(wc,3,"COGS",cogs)
    _xl_sub(wc,4,"Gross Profit","=B2-B3"); _xl_given(wc,5,"SG&A",sga)
    _xl_sub(wc,6,"EBITDA","=B4-B5"); _xl_given(wc,7,"D&A",da)
    _xl_sub(wc,8,"EBIT","=B6-B7"); _xl_given(wc,9,"Interest Expense",interest)
    _xl_sub(wc,10,"EBT","=B8-B9"); _xl_given(wc,11,"Tax Rate",TAX)
    wc["B11"].number_format="0%"
    _xl_fml(wc,12,"Tax Expense","=B10*B11"); _xl_sub(wc,13,"Net Income","=B10-B12")
    _xl_fml(wc,14,"NI Margin","=B13/B2"); wc["B14"].number_format="0.0%"

    # IS — Practice
    wp = wb.create_sheet("IS — Practice"); _cw(wp, [("A",28),("B",16)])
    _xl_hdr_row(wp,1,f"Income Statement — {nm}","($M)")
    _xl_prac(wp,2,"Revenue",R); _xl_prac(wp,3,"COGS",cogs)
    _xl_prac(wp,4,"Gross Profit"); _xl_prac(wp,5,"SG&A",sga)
    _xl_prac(wp,6,"EBITDA"); _xl_prac(wp,7,"D&A",da)
    _xl_prac(wp,8,"EBIT"); _xl_prac(wp,9,"Interest Expense",interest)
    _xl_prac(wp,10,"EBT"); _xl_prac(wp,11,"Tax Rate",TAX)
    wp["B11"].number_format="0%"
    _xl_prac(wp,12,"Tax Expense"); _xl_prac(wp,13,"Net Income"); _xl_prac(wp,14,"NI Margin")

    # BS — Complete
    wc2 = wb.create_sheet("BS — Complete"); _cw(wc2,[("A",28),("B",16)])
    _xl_hdr_row(wc2,1,f"Balance Sheet — {nm}","($M)")
    _xl_given(wc2,2,"Cash",cash); _xl_given(wc2,3,"Accounts Receivable",ar)
    _xl_given(wc2,4,"Inventory",inv); _xl_sub(wc2,5,"Current Assets","=B2+B3+B4")
    _xl_given(wc2,6,"PP&E, net",ppe); _xl_given(wc2,7,"Intangibles",intang)
    _xl_sub(wc2,8,"Total Assets","=B5+B6+B7")
    _xl_given(wc2,10,"Accounts Payable",ap); _xl_given(wc2,11,"Short-term Debt",st_debt)
    _xl_given(wc2,12,"Long-term Debt",lt_debt); _xl_sub(wc2,13,"Total Liabilities","=B10+B11+B12")
    _xl_given(wc2,15,"Common Stock",com_stock); _xl_given(wc2,16,"Retained Earnings",ret_earn)
    _xl_sub(wc2,17,"Total Equity","=B15+B16"); _xl_sub(wc2,18,"Total L&E","=B13+B17")
    _xl_fml(wc2,20,"Balance Check (must = 0)","=B8-B18")

    # BS — Practice
    wp2 = wb.create_sheet("BS — Practice"); _cw(wp2,[("A",28),("B",16)])
    _xl_hdr_row(wp2,1,f"Balance Sheet — {nm}","($M)")
    _xl_prac(wp2,2,"Cash",cash); _xl_prac(wp2,3,"Accounts Receivable",ar)
    _xl_prac(wp2,4,"Inventory",inv); _xl_prac(wp2,5,"Current Assets")
    _xl_prac(wp2,6,"PP&E, net",ppe); _xl_prac(wp2,7,"Intangibles",intang)
    _xl_prac(wp2,8,"Total Assets"); _xl_prac(wp2,10,"Accounts Payable",ap)
    _xl_prac(wp2,11,"Short-term Debt",st_debt); _xl_prac(wp2,12,"Long-term Debt",lt_debt)
    _xl_prac(wp2,13,"Total Liabilities"); _xl_prac(wp2,15,"Common Stock",com_stock)
    _xl_prac(wp2,16,"Retained Earnings",ret_earn); _xl_prac(wp2,17,"Total Equity")
    _xl_prac(wp2,18,"Total L&E"); _xl_prac(wp2,20,"Balance Check (must = 0)")

    # CFS — Complete
    wc3 = wb.create_sheet("CFS — Complete"); _cw(wc3,[("A",32),("B",16)])
    _xl_hdr_row(wc3,1,f"Cash Flow Statement — {nm}","($M)")
    _xl_sec(wc3,2,"— Operating Activities —")
    _xl_fml(wc3,3,"Net Income","='IS — Complete'!B13")
    _xl_fml(wc3,4,"Add: D&A","='IS — Complete'!B7")
    _xl_given(wc3,5,"Δ Accounts Receivable",0); _xl_given(wc3,6,"Δ Inventory",0)
    _xl_given(wc3,7,"Δ Accounts Payable",0)
    _xl_sub(wc3,8,"Cash from Operations","=B3+B4+B5+B6+B7")
    _xl_sec(wc3,10,"— Investing Activities —")
    _xl_given(wc3,11,"Capex",-capex); _xl_sub(wc3,12,"Cash from Investing","=B11")
    _xl_sec(wc3,14,"— Financing Activities —")
    _xl_given(wc3,15,"Dividends Paid",-divs); _xl_sub(wc3,16,"Cash from Financing","=B15")
    _xl_sub(wc3,18,"Ending Cash","='BS — Complete'!B2+B8+B12+B16")

    # CFS — Practice
    wp3 = wb.create_sheet("CFS — Practice"); _cw(wp3,[("A",32),("B",16)])
    _xl_hdr_row(wp3,1,f"Cash Flow Statement — {nm}","($M)")
    _xl_sec(wp3,2,"— Operating Activities —")
    _xl_prac(wp3,3,"Net Income  ← link from IS tab")
    _xl_prac(wp3,4,"Add: D&A  ← link from IS tab")
    _xl_prac(wp3,5,"Δ Accounts Receivable",0); _xl_prac(wp3,6,"Δ Inventory",0)
    _xl_prac(wp3,7,"Δ Accounts Payable",0); _xl_prac(wp3,8,"Cash from Operations")
    _xl_sec(wp3,10,"— Investing Activities —")
    _xl_prac(wp3,11,"Capex",-capex); _xl_prac(wp3,12,"Cash from Investing")
    _xl_sec(wp3,14,"— Financing Activities —")
    _xl_prac(wp3,15,"Dividends Paid",-divs); _xl_prac(wp3,16,"Cash from Financing")
    _xl_prac(wp3,18,"Ending Cash  ← Beginning Cash + Ops + Investing + Financing")


def _make_dcf_excel(wb, co):
    R = co["rev_n"]; E = co["ebitda_n"]; S = co["shares_n"]
    g = co.get("rev_growth",0.05); wc_r = co.get("wacc",0.0935)
    tg = co.get("terminal_g",0.025); TAX = 0.30
    net_debt = round(E*1.05/5)*5; nm = co["name"]

    # Complete
    wc = wb.create_sheet("DCF — Complete")
    _cw(wc,[("A",30),("B",13),("C",12),("D",12),("E",12),("F",12),("G",12)])
    _xl_hdr_row(wc,1,f"DCF — {nm}","Assumptions","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    _xl_sec(wc,3,"Assumptions (col B)")
    _xl_given(wc,4,"Base Revenue ($M)",R); _xl_given(wc,5,"Revenue Growth",g)
    wc["B5"].number_format="0.0%"
    _xl_given(wc,6,"EBITDA Margin",round(E/R,4)); wc["B6"].number_format="0.0%"
    _xl_given(wc,7,"D&A % Revenue",0.06); wc["B7"].number_format="0.0%"
    _xl_given(wc,8,"Capex % Revenue",0.07); wc["B8"].number_format="0.0%"
    _xl_given(wc,9,"Tax Rate",TAX); wc["B9"].number_format="0%"
    _xl_given(wc,10,"WACC",wc_r); wc["B10"].number_format="0.00%"
    _xl_given(wc,11,"Terminal Growth (g)",tg); wc["B11"].number_format="0.0%"
    _xl_given(wc,12,"Net Debt ($M)",net_debt); _xl_given(wc,13,"Shares Out. (M)",S)
    _xl_sec(wc,15,"5-Year Projections (cols C–G)")
    for lbl,r in [("Revenue ($M)",16),("EBITDA ($M)",17),("Free Cash Flow ($M)",18),
                  ("Discount Factor",19),("PV of FCF ($M)",20)]:
        wc[f"A{r}"].value = lbl; wc[f"A{r}"].font = _BOLD_FONT
    for yr,col in enumerate(["C","D","E","F","G"],1):
        wc[f"{col}16"].value=f"=$B$4*(1+$B$5)^{yr}"; wc[f"{col}16"].font=_GREEN_FONT
        wc[f"{col}17"].value=f"={col}16*$B$6"; wc[f"{col}17"].font=_GREEN_FONT
        wc[f"{col}18"].value=f"={col}17*(1-$B$9)-{col}16*$B$8"; wc[f"{col}18"].font=_GREEN_FONT
        wc[f"{col}19"].value=f"=1/(1+$B$10)^{yr}"; wc[f"{col}19"].font=_GREEN_FONT
        wc[f"{col}20"].value=f"={col}18*{col}19"; wc[f"{col}20"].font=_GREEN_FONT
    _xl_sec(wc,22,"Terminal Value & Bridge (col B)")
    _xl_fml(wc,23,"Year 5 FCF ($M)","=G18")
    _xl_fml(wc,24,"Terminal Value ($M)","=B23*(1+B11)/(B10-B11)")
    _xl_fml(wc,25,"PV of Terminal Value","=B24/(1+B10)^5")
    _xl_fml(wc,26,"Sum of PV(FCFs)","=SUM(C20:G20)")
    _xl_sub(wc,27,"Enterprise Value ($M)","=B25+B26")
    _xl_given(wc,28,"Net Debt ($M)",net_debt)
    _xl_sub(wc,29,"Equity Value ($M)","=B27-B28")
    _xl_sub(wc,30,"Implied Share Price","=B29/B13")
    wc["B30"].number_format='"$"#,##0.00'

    # Practice
    wp = wb.create_sheet("DCF — Practice")
    _cw(wp,[("A",30),("B",13),("C",12),("D",12),("E",12),("F",12),("G",12)])
    _xl_hdr_row(wp,1,f"DCF — {nm}","Assumptions","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    _xl_sec(wp,3,"Assumptions (col B)")
    _xl_prac(wp,4,"Base Revenue ($M)",R); _xl_prac(wp,5,"Revenue Growth",g)
    _xl_prac(wp,6,"EBITDA Margin",round(E/R,4)); _xl_prac(wp,7,"D&A % Revenue",0.06)
    _xl_prac(wp,8,"Capex % Revenue",0.07); _xl_prac(wp,9,"Tax Rate",TAX)
    _xl_prac(wp,10,"WACC",wc_r); _xl_prac(wp,11,"Terminal Growth (g)",tg)
    _xl_prac(wp,12,"Net Debt ($M)",net_debt); _xl_prac(wp,13,"Shares Out. (M)",S)
    _xl_sec(wp,15,"5-Year Projections (cols C–G)")
    for lbl,r in [("Revenue ($M)",16),("EBITDA ($M)",17),("Free Cash Flow ($M)",18),
                  ("Discount Factor",19),("PV of FCF ($M)",20)]:
        wp[f"A{r}"].value = lbl; wp[f"A{r}"].font = _BOLD_FONT
        for col in ["C","D","E","F","G"]:
            wp[f"{col}{r}"].fill = _YELLOW_FILL
    _xl_sec(wp,22,"Terminal Value & Bridge (col B)")
    for lbl,r,val in [("Year 5 FCF ($M)",23,None),("Terminal Value ($M)",24,None),
                       ("PV of Terminal Value",25,None),("Sum of PV(FCFs)",26,None),
                       ("Enterprise Value ($M)",27,None),("Net Debt ($M)",28,net_debt),
                       ("Equity Value ($M)",29,None),("Implied Share Price",30,None)]:
        if val is not None:
            _xl_prac(wp,r,lbl,val)
        else:
            _xl_prac(wp,r,lbl)


def _make_comps_excel(wb, co):
    E = co["ebitda_n"]; S = co["shares_n"]; nm = co["name"]
    net_debt = round(E*1.05/5)*5
    peers = [("TechAlpha",1104,120,480,65,1202),("DataCore",810,100,420,50,810),
             ("CloudSystems",1260,120,400,59,1239),("InfoPro",858,110,510,54,853)]

    # Complete
    wc = wb.create_sheet("Comps — Complete")
    _cw(wc,[("A",20),("B",12),("C",12),("D",12),("E",13),("F",12),("G",12),("H",12)])
    _xl_hdr_row(wc,1,"Company","EV ($M)","EBITDA ($M)","Revenue ($M)","Net Inc ($M)","Mkt Cap ($M)","EV/EBITDA","EV/Revenue")
    for i,(pnm,ev,ebitda,rev,ni,mktcap) in enumerate(peers,2):
        wc[f"A{i}"].value=pnm
        for col,val in zip(["B","C","D","E","F"],[ev,ebitda,rev,ni,mktcap]):
            wc[f"{col}{i}"].value=val; wc[f"{col}{i}"].fill=_BLUE_FILL; wc[f"{col}{i}"].font=_BLUE_FONT
        wc[f"G{i}"].value=f"=B{i}/C{i}"; wc[f"G{i}"].font=_GREEN_FONT; wc[f"G{i}"].number_format='0.0"×"'
        wc[f"H{i}"].value=f"=B{i}/D{i}"; wc[f"H{i}"].font=_GREEN_FONT; wc[f"H{i}"].number_format='0.0"×"'
    _xl_sec(wc,7,"Statistics")
    for lbl,r,pct in [("25th Pctile",8,0.25),("Median",9,0.5),("75th Pctile",10,0.75)]:
        wc[f"A{r}"].value=lbl; wc[f"A{r}"].font=_BOLD_BL
        wc[f"G{r}"].value=f"=MEDIAN(G2:G5)" if pct==0.5 else f"=PERCENTILE(G2:G5,{pct})"
        wc[f"G{r}"].font=_BOLD_BL; wc[f"G{r}"].fill=_SUB_FILL; wc[f"G{r}"].number_format='0.0"×"'
        wc[f"H{r}"].value=f"=MEDIAN(H2:H5)" if pct==0.5 else f"=PERCENTILE(H2:H5,{pct})"
        wc[f"H{r}"].font=_BOLD_BL; wc[f"H{r}"].fill=_SUB_FILL; wc[f"H{r}"].number_format='0.0"×"'
    _xl_sec(wc,12,f"Implied Valuation — {nm}")
    _xl_given(wc,13,f"{nm} EBITDA ($M)",E); _xl_fml(wc,14,"Median EV/EBITDA","=G9")
    _xl_sub(wc,15,"Implied EV ($M)","=B13*B14"); _xl_given(wc,16,"Net Debt ($M)",net_debt)
    _xl_sub(wc,17,"Equity Value ($M)","=B15-B16"); _xl_given(wc,18,"Shares Out. (M)",S)
    _xl_sub(wc,19,"Implied Share Price","=B17/B18"); wc["B19"].number_format='"$"#,##0.00'

    # Practice
    wp = wb.create_sheet("Comps — Practice")
    _cw(wp,[("A",20),("B",12),("C",12),("D",12),("E",13),("F",12),("G",12),("H",12)])
    _xl_hdr_row(wp,1,"Company","EV ($M)","EBITDA ($M)","Revenue ($M)","Net Inc ($M)","Mkt Cap ($M)","EV/EBITDA","EV/Revenue")
    for i,(pnm,ev,ebitda,rev,ni,mktcap) in enumerate(peers,2):
        wp[f"A{i}"].value=pnm
        for col,val in zip(["B","C","D","E","F"],[ev,ebitda,rev,ni,mktcap]):
            wp[f"{col}{i}"].value=val; wp[f"{col}{i}"].fill=_BLUE_FILL; wp[f"{col}{i}"].font=_BLUE_FONT
        wp[f"G{i}"].fill=_YELLOW_FILL; wp[f"H{i}"].fill=_YELLOW_FILL
    _xl_sec(wp,7,"Statistics")
    for lbl,r in [("25th Pctile",8),("Median",9),("75th Pctile",10)]:
        wp[f"A{r}"].value=lbl; wp[f"A{r}"].font=_BOLD_BL
        wp[f"G{r}"].fill=_YELLOW_FILL; wp[f"H{r}"].fill=_YELLOW_FILL
    _xl_sec(wp,12,f"Implied Valuation — {nm}")
    _xl_prac(wp,13,f"{nm} EBITDA ($M)",E); _xl_prac(wp,14,"Median EV/EBITDA")
    _xl_prac(wp,15,"Implied EV ($M)"); _xl_prac(wp,16,"Net Debt ($M)",net_debt)
    _xl_prac(wp,17,"Equity Value ($M)"); _xl_prac(wp,18,"Shares Out. (M)",S)
    _xl_prac(wp,19,"Implied Share Price")


def _make_precedent_excel(wb, co):
    E = co["ebitda_n"]; S = co["shares_n"]; nm = co["name"]
    net_debt = round(E*1.05/5)*5
    deals = [("AcquireCo / TechAlpha '22",1320,120,975,2022),
             ("MegaCorp / DataCore '21",980,100,809,2021),
             ("GlobalTech / InfoPro '23",1045,110,856,2023)]

    # Complete
    wc = wb.create_sheet("Precedent — Complete")
    _cw(wc,[("A",28),("B",8),("C",13),("D",14),("E",13),("F",16)])
    _xl_hdr_row(wc,1,"Deal","Year","Deal EV ($M)","LTM EBITDA ($M)","EV/EBITDA","Pre-deal Mkt Cap ($M)")
    for i,(dnm,ev,ebitda,predeal,yr) in enumerate(deals,2):
        wc[f"A{i}"].value=dnm
        for col,val in zip(["B","C","D","F"],[yr,ev,ebitda,predeal]):
            wc[f"{col}{i}"].value=val; wc[f"{col}{i}"].fill=_BLUE_FILL; wc[f"{col}{i}"].font=_BLUE_FONT
        wc[f"E{i}"].value=f"=C{i}/D{i}"; wc[f"E{i}"].font=_GREEN_FONT; wc[f"E{i}"].number_format='0.0"×"'
    _xl_sec(wc,6,"Statistics")
    wc["A7"].value="Median EV/EBITDA"; wc["A7"].font=_BOLD_BL
    wc["E7"].value="=MEDIAN(E2:E4)"; wc["E7"].font=_BOLD_BL
    wc["E7"].fill=_SUB_FILL; wc["E7"].number_format='0.0"×"'
    _xl_sec(wc,9,f"Implied Valuation — {nm}")
    _xl_given(wc,10,f"{nm} LTM EBITDA ($M)",E); _xl_fml(wc,11,"Median Deal Multiple","=E7")
    _xl_sub(wc,12,"Implied EV ($M)","=B10*B11"); _xl_given(wc,13,"Net Debt ($M)",net_debt)
    _xl_sub(wc,14,"Equity Value ($M)","=B12-B13"); _xl_given(wc,15,"Shares Out. (M)",S)
    _xl_sub(wc,16,"Implied Share Price","=B14/B15"); wc["B16"].number_format='"$"#,##0.00'

    # Practice
    wp = wb.create_sheet("Precedent — Practice")
    _cw(wp,[("A",28),("B",8),("C",13),("D",14),("E",13),("F",16)])
    _xl_hdr_row(wp,1,"Deal","Year","Deal EV ($M)","LTM EBITDA ($M)","EV/EBITDA","Pre-deal Mkt Cap ($M)")
    for i,(dnm,ev,ebitda,predeal,yr) in enumerate(deals,2):
        wp[f"A{i}"].value=dnm
        for col,val in zip(["B","C","D","F"],[yr,ev,ebitda,predeal]):
            wp[f"{col}{i}"].value=val; wp[f"{col}{i}"].fill=_BLUE_FILL; wp[f"{col}{i}"].font=_BLUE_FONT
        wp[f"E{i}"].fill=_YELLOW_FILL
    _xl_sec(wp,6,"Statistics")
    wp["A7"].value="Median EV/EBITDA"; wp["A7"].font=_BOLD_BL; wp["E7"].fill=_YELLOW_FILL
    _xl_sec(wp,9,f"Implied Valuation — {nm}")
    _xl_prac(wp,10,f"{nm} LTM EBITDA ($M)",E); _xl_prac(wp,11,"Median Deal Multiple")
    _xl_prac(wp,12,"Implied EV ($M)"); _xl_prac(wp,13,"Net Debt ($M)",net_debt)
    _xl_prac(wp,14,"Equity Value ($M)"); _xl_prac(wp,15,"Shares Out. (M)",S)
    _xl_prac(wp,16,"Implied Share Price")


def _make_lbo_excel(wb, co):
    E = co["ebitda_n"]; nm = co["name"]; TAX = 0.30
    capex = round(co["rev_n"]*0.07,1); g = co.get("rev_growth",0.05)
    ir = 0.07; entry_m = 9.0

    # Complete
    wc = wb.create_sheet("LBO — Complete")
    _cw(wc,[("A",28),("B",13),("C",12),("D",12),("E",12),("F",12),("G",12)])
    _xl_hdr_row(wc,1,f"LBO — {nm}","Entry","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    _xl_sec(wc,2,"Sources & Uses")
    _xl_given(wc,3,"LTM EBITDA ($M)",E); _xl_given(wc,4,"Entry EV/EBITDA",entry_m)
    _xl_sub(wc,5,"Entry EV ($M)","=B3*B4")
    _xl_fml(wc,6,"Debt (60%)","=B5*0.60"); _xl_fml(wc,7,"PE Equity (40%)","=B5*0.40")
    _xl_fml(wc,8,"Sources Check","=B6+B7")
    _xl_sec(wc,10,"Debt Schedule")
    _xl_hdr_row(wc,11,"","Entry","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    for lbl,r in [("Beginning Debt",12),("Interest Expense",13),("EBITDA",14),
                  ("FCF for Paydown",15),("Ending Debt",16)]:
        wc[f"A{r}"].value=lbl
    # Entry col (B): show initial debt
    wc["B12"].value="=B6"; wc["B12"].font=_GREEN_FONT
    for yr,col in enumerate(["C","D","E","F","G"],1):
        prev = "B" if yr==1 else chr(ord(col)-1)
        wc[f"{col}12"].value=f"={prev}{'6' if yr==1 else '16'}"; wc[f"{col}12"].font=_GREEN_FONT
        wc[f"{col}13"].value=f"={col}12*{ir}"; wc[f"{col}13"].font=_GREEN_FONT
        wc[f"{col}14"].value=f"=B3*(1+{g})^{yr}"; wc[f"{col}14"].font=_GREEN_FONT
        wc[f"{col}15"].value=f"={col}14*(1-{TAX})-{capex}"; wc[f"{col}15"].font=_GREEN_FONT
        wc[f"{col}16"].value=f"={col}12-{col}15"; wc[f"{col}16"].font=_GREEN_FONT
    _xl_sec(wc,18,"Exit & Returns")
    _xl_given(wc,19,"Exit EV/EBITDA",entry_m)
    _xl_fml(wc,20,"Year 5 EBITDA ($M)",f"=B3*(1+{g})^5")
    _xl_sub(wc,21,"Exit EV ($M)","=B19*B20"); _xl_fml(wc,22,"Remaining Debt ($M)","=G16")
    _xl_sub(wc,23,"Exit Equity ($M)","=B21-B22"); _xl_fml(wc,24,"Entry Equity ($M)","=B7")
    _xl_sub(wc,25,"MOIC","=B23/B24"); wc["B25"].number_format='0.00"×"'

    # Practice
    wp = wb.create_sheet("LBO — Practice")
    _cw(wp,[("A",28),("B",13),("C",12),("D",12),("E",12),("F",12),("G",12)])
    _xl_hdr_row(wp,1,f"LBO — {nm}","Entry","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    _xl_sec(wp,2,"Sources & Uses")
    _xl_prac(wp,3,"LTM EBITDA ($M)",E); _xl_prac(wp,4,"Entry EV/EBITDA",entry_m)
    _xl_prac(wp,5,"Entry EV ($M)"); _xl_prac(wp,6,"Debt (60%)")
    _xl_prac(wp,7,"PE Equity (40%)"); _xl_prac(wp,8,"Sources Check")
    _xl_sec(wp,10,"Debt Schedule")
    _xl_hdr_row(wp,11,"","Entry","Yr 1","Yr 2","Yr 3","Yr 4","Yr 5")
    for lbl,r in [("Beginning Debt",12),("Interest Expense",13),("EBITDA",14),
                  ("FCF for Paydown",15),("Ending Debt",16)]:
        wp[f"A{r}"].value=lbl
        for col in ["B","C","D","E","F","G"]:
            wp[f"{col}{r}"].fill=_YELLOW_FILL
    _xl_sec(wp,18,"Exit & Returns")
    _xl_prac(wp,19,"Exit EV/EBITDA",entry_m); _xl_prac(wp,20,"Year 5 EBITDA ($M)")
    _xl_prac(wp,21,"Exit EV ($M)"); _xl_prac(wp,22,"Remaining Debt ($M)")
    _xl_prac(wp,23,"Exit Equity ($M)"); _xl_prac(wp,24,"Entry Equity ($M)")
    _xl_prac(wp,25,"MOIC")


def _make_merger_excel(wb, co):
    N = co["ni_n"]; P = co["price_n"]; S = co["shares_n"]; nm = co["name"]
    tgt_price = round(P*0.60,2); tgt_shares = round(S*0.40,1)
    tgt_ni = round(N*0.35,1); syn = round(N*0.08,1); amort = round(N*0.04,1)

    # Complete
    wc = wb.create_sheet("Merger — Complete"); _cw(wc,[("A",32),("B",14)])
    _xl_sec(wc,1,"Deal Assumptions")
    _xl_given(wc,2,"Target Market Price",tgt_price); wc["B2"].number_format='"$"#,##0.00'
    _xl_given(wc,3,"Acquisition Premium",0.25); wc["B3"].number_format="0%"
    _xl_sub(wc,4,"Offer Price / Share","=B2*(1+B3)"); wc["B4"].number_format='"$"#,##0.00'
    _xl_given(wc,5,"Target Shares Out. (M)",tgt_shares)
    _xl_sub(wc,6,"Total Deal Value ($M)","=B4*B5")
    _xl_given(wc,7,f"{nm} Share Price",P); wc["B7"].number_format='"$"#,##0.00'
    _xl_fml(wc,8,"New Shares Issued (M)","=B6/B7")
    _xl_fml(wc,9,"Dilution %",f"=B8/(B8+{S})"); wc["B9"].number_format="0.0%"
    _xl_sec(wc,11,"Pro Forma Income Statement")
    _xl_given(wc,12,f"Acquirer NI — {nm} ($M)",N); _xl_given(wc,13,"Target NI ($M)",tgt_ni)
    _xl_given(wc,14,"After-tax Synergies ($M)",syn)
    _xl_given(wc,15,"Intangibles Amortization ($M)",-amort)
    _xl_sub(wc,16,"Pro Forma NI ($M)","=SUM(B12:B15)")
    _xl_sec(wc,18,"EPS & Accretion / Dilution")
    _xl_given(wc,19,"Standalone Shares (M)",S)
    _xl_fml(wc,20,"New Shares Issued (M)","=B8"); _xl_sub(wc,21,"Pro Forma Shares (M)","=B19+B20")
    _xl_fml(wc,22,"Standalone EPS",f"=B12/B19"); wc["B22"].number_format='"$"#,##0.00'
    _xl_fml(wc,23,"Pro Forma EPS","=B16/B21"); wc["B23"].number_format='"$"#,##0.00'
    _xl_sub(wc,24,"Accretion / (Dilution)","=(B23-B22)/B22"); wc["B24"].number_format="0.0%"

    # Practice
    wp = wb.create_sheet("Merger — Practice"); _cw(wp,[("A",32),("B",14)])
    _xl_sec(wp,1,"Deal Assumptions")
    _xl_prac(wp,2,"Target Market Price",tgt_price); _xl_prac(wp,3,"Acquisition Premium",0.25)
    _xl_prac(wp,4,"Offer Price / Share"); _xl_prac(wp,5,"Target Shares Out. (M)",tgt_shares)
    _xl_prac(wp,6,"Total Deal Value ($M)"); _xl_prac(wp,7,f"{nm} Share Price",P)
    _xl_prac(wp,8,"New Shares Issued (M)"); _xl_prac(wp,9,"Dilution %")
    _xl_sec(wp,11,"Pro Forma Income Statement")
    _xl_prac(wp,12,f"Acquirer NI — {nm} ($M)",N); _xl_prac(wp,13,"Target NI ($M)",tgt_ni)
    _xl_prac(wp,14,"After-tax Synergies ($M)",syn)
    _xl_prac(wp,15,"Intangibles Amortization ($M)",-amort)
    _xl_prac(wp,16,"Pro Forma NI ($M)")
    _xl_sec(wp,18,"EPS & Accretion / Dilution")
    _xl_prac(wp,19,"Standalone Shares (M)",S); _xl_prac(wp,20,"New Shares Issued (M)")
    _xl_prac(wp,21,"Pro Forma Shares (M)"); _xl_prac(wp,22,"Standalone EPS")
    _xl_prac(wp,23,"Pro Forma EPS"); _xl_prac(wp,24,"Accretion / (Dilution)")


def _make_budget_excel(wb, co):
    R = co["rev_n"]; E = co["ebitda_n"]; nm = co["name"]
    cogs = round(R*0.60,1); sga = round(R*0.40-E,1)
    UB,UA = 12_000,11_500
    rb = round(R*0.24,1); ra = round(rb*0.942,1)
    cb = round(cogs*0.24,1); ca = round(cb*0.976,1)
    sb = round(sga*0.25,1); sa = round(sb*1.075,1)
    pb = round(rb*1e6/UB); pa = round(ra*1e6/UA)

    # Complete
    wc = wb.create_sheet("BvA — Complete")
    _cw(wc,[("A",20),("B",12),("C",12),("D",13),("E",12),("F",6)])
    _xl_hdr_row(wc,1,f"Budget vs. Actual — {nm} Q1","Budget ($M)","Actual ($M)","Variance $","Variance %","F/U")
    data = [(2,"Revenue",rb,ra,True),(3,"COGS",cb,ca,False),
            (4,"Gross Profit",None,None,True),(5,"SG&A",sb,sa,False),
            (6,"EBITDA",None,None,True)]
    for r,lbl,bv,av,fav in data:
        wc[f"A{r}"].value=lbl
        if bv is not None:
            wc[f"B{r}"].value=bv; wc[f"B{r}"].fill=_BLUE_FILL; wc[f"B{r}"].font=_BLUE_FONT
            wc[f"C{r}"].value=av; wc[f"C{r}"].fill=_BLUE_FILL; wc[f"C{r}"].font=_BLUE_FONT
        else:
            gp_fml = ("=B2-B3","=C2-C3") if lbl=="Gross Profit" else ("=B4-B5","=C4-C5")
            wc[f"B{r}"].value=gp_fml[0]; wc[f"B{r}"].font=_GREEN_FONT
            wc[f"C{r}"].value=gp_fml[1]; wc[f"C{r}"].font=_GREEN_FONT
        wc[f"D{r}"].value=f"=C{r}-B{r}"; wc[f"D{r}"].font=_GREEN_FONT
        wc[f"E{r}"].value=f"=D{r}/B{r}"; wc[f"E{r}"].font=_GREEN_FONT
        wc[f"E{r}"].number_format="0.0%"
        flag=f'=IF(D{r}>=0,"F","U")' if fav else f'=IF(D{r}<=0,"F","U")'
        wc[f"F{r}"].value=flag; wc[f"F{r}"].font=_GREEN_FONT
    _xl_sec(wc,8,"Volume / Price Decomposition")
    _xl_given(wc,9,"Budget Units",UB); _xl_given(wc,10,"Actual Units",UA)
    _xl_given(wc,11,"Budget Price/Unit ($)",pb); _xl_given(wc,12,"Actual Price/Unit ($)",pa)
    _xl_fml(wc,13,"Volume Effect ($M)","=(B10-B9)*B11/1000000")
    _xl_fml(wc,14,"Price Effect ($M)","=B10*(B12-B11)/1000000")
    _xl_sub(wc,15,"Check (= Revenue Variance)","=B13+B14")

    # Practice
    wp = wb.create_sheet("BvA — Practice")
    _cw(wp,[("A",20),("B",12),("C",12),("D",13),("E",12),("F",6)])
    _xl_hdr_row(wp,1,f"Budget vs. Actual — {nm} Q1","Budget ($M)","Actual ($M)","Variance $","Variance %","F/U")
    for r,lbl,bv,av,_ in data:
        wp[f"A{r}"].value=lbl
        if bv is not None:
            wp[f"B{r}"].value=bv; wp[f"B{r}"].fill=_BLUE_FILL; wp[f"B{r}"].font=_BLUE_FONT
            wp[f"C{r}"].value=av; wp[f"C{r}"].fill=_BLUE_FILL; wp[f"C{r}"].font=_BLUE_FONT
        else:
            wp[f"B{r}"].fill=_YELLOW_FILL; wp[f"C{r}"].fill=_YELLOW_FILL
        for col in ["D","E","F"]:
            wp[f"{col}{r}"].fill=_YELLOW_FILL
    _xl_sec(wp,8,"Volume / Price Decomposition")
    _xl_prac(wp,9,"Budget Units",UB); _xl_prac(wp,10,"Actual Units",UA)
    _xl_prac(wp,11,"Budget Price/Unit ($)",pb); _xl_prac(wp,12,"Actual Price/Unit ($)",pa)
    _xl_prac(wp,13,"Volume Effect ($M)"); _xl_prac(wp,14,"Price Effect ($M)")
    _xl_prac(wp,15,"Check (= Revenue Variance)")


def make_model_excel(model_key, co):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    {"3stmt": _make_3stmt_excel, "dcf": _make_dcf_excel, "comps": _make_comps_excel,
     "precedent": _make_precedent_excel, "lbo": _make_lbo_excel,
     "merger": _make_merger_excel, "budget": _make_budget_excel
    }[model_key](wb, co)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _fs(ws, r, lbl, val=None, ind=0, bold=False, ul=False, hdr=False, note=False):
    """Write one row of a formatted financial statement."""
    ws[f"A{r}"].value = ("    " * ind) + lbl
    if hdr:
        ws[f"A{r}"].font = Font(bold=True, size=9)
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="E2E8F0")
    elif note:
        ws[f"A{r}"].font = Font(size=8, italic=True, color="94A3B8")
    else:
        ws[f"A{r}"].font = Font(bold=bold, size=9)
    if val is not None:
        ws[f"B{r}"].value = val
        ws[f"B{r}"].number_format = '#,##0.0'
        ws[f"B{r}"].alignment = _RIGHT_ALIGN
        ws[f"B{r}"].font = Font(bold=bold or hdr, size=9, color="94A3B8" if note else "000000")
        if ul:
            ws[f"B{r}"].border = Border(bottom=Side(style="thin", color="374151"))
        if hdr:
            ws[f"B{r}"].fill = PatternFill("solid", fgColor="E2E8F0")


def _fs_title(ws, r, text, sub=False):
    ws[f"A{r}"].value = text
    ws[f"A{r}"].font = Font(bold=not sub, size=11 if not sub else 9,
                             color="1E3A5F" if not sub else "6B7280",
                             italic=sub)


def _adv_header(ws, r, nm):
    """Instruction banner at top of every Advanced sheet."""
    ws[f"A{r}"].value = (
        "ADVANCED MODELING  —  Read the full financial statements below. "
        "Extract only the figures your model requires, then build the model entirely from scratch on a new sheet.")
    ws[f"A{r}"].font = Font(bold=True, size=9, color="FFFFFF")
    ws[f"A{r}"].fill = PatternFill("solid", fgColor="7C3AED")
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 30


def _make_adv_3stmt(wb, co):
    R=co["rev_n"]; E=co["ebitda_n"]; N=co["ni_n"]; nm=co["name"]; S=co["shares_n"]
    TAX=0.30
    cogs=round(R*.60,1); gp=round(R-cogs,1); sga=round(R*.40-E,1); da=round(R*.06,1)
    ebit=E-da; ebt=round(N/(1-TAX),1); interest_net=round(ebit-ebt,1); tax=round(ebt*TAX,1)
    capex=round(R*.07,1); divs=round(N*.25,1); nd=round(E*1.05/5)*5
    # IS sub-items
    prod_rev=round(R*.70,1); svc_rev=round(R-prod_rev,1)
    cogs_prod=round(cogs*.80,1); cogs_svc=round(cogs-cogs_prod,1)
    rd=round(sga*.40,1); sm=round(sga*.35,1); ga=round(sga-rd-sm,1)
    int_inc=round(ebt*.02,1); int_exp=round(interest_net+int_inc,1)
    # BS (real company has extra line items)
    cash_cce=round(R*.13,1); cash_stinv=round(R*.02,1); cash_tot=cash_cce+cash_stinv
    ar=round(R*.12,1); inv=round(R*.08,1)
    prepaid=round(R*.017,1); tax_rec=round(R*.006,1)
    curr_a=round(cash_tot+ar+inv+prepaid+tax_rec,1)
    ppe=round(R*.40,1); goodwill=round(R*.03,1); intang=round(R*.10,1); other_lta=round(R*.01,1)
    tot_a=round(curr_a+ppe+goodwill+intang+other_lta,1)
    ap=round(R*.06,1); std=round(nd*.15/5)*5; ltd=nd+cash_tot-std
    accrued=round(R*.016,1); oth_acc=round(R*.008,1); tax_pay=round(R*.005,1)
    curr_l=round(ap+std+accrued+oth_acc+tax_pay,1)
    def_rev=round(R*.010,1); oth_ltl=round(R*.006,1); tot_l=round(curr_l+ltd+def_rev+oth_ltl,1)
    tot_eq=round(tot_a-tot_l,1); aoci=round(-R*.004,1)
    cs=round((tot_eq-aoci)*.52,1); re=round(tot_eq-aoci-cs,1)
    # CFS extra items
    sbc=round(N*.10,1); d_ar=round(-R*.007,1); d_inv=round(-R*.004,1)
    d_ap=round(R*.004,1); oth_wc=round(R*.002,1)
    cfo=round(N+da+sbc+d_ar+d_inv+d_ap+oth_wc,1)
    cfi=round(-capex-R*.04+R*.016,1)
    debt_pay=round(-nd*.04,1); sop=round(N*.04,1); tax_sop=round(-N*.027,1)
    cff=round(-divs+debt_pay+sop+tax_sop,1)
    net_chg=round(cfo+cfi+cff,1); beg_cash=round(cash_cce-net_chg,1)

    ws=wb.create_sheet("Advanced Practice"); _cw(ws,[("A",50),("B",14)])
    r=1; _adv_header(ws,r,nm); r+=2
    # ── IS ──────────────────────────────────────────────────────────
    _fs_title(ws,r,nm); r+=1
    _fs_title(ws,r,"CONSOLIDATED STATEMENT OF OPERATIONS"); r+=1
    _fs_title(ws,r,"(Amounts in millions of U.S. dollars, except per share data — Fiscal Year 2024)",sub=True); r+=2
    _fs(ws,r,"Revenues",hdr=True); r+=1
    _fs(ws,r,"Product revenue",prod_rev,ind=1); r+=1
    _fs(ws,r,"Service and subscription revenue",svc_rev,ind=1,ul=True); r+=1
    _fs(ws,r,"Total revenues",R,bold=True); r+=2
    _fs(ws,r,"Costs and expenses",hdr=True); r+=1
    _fs(ws,r,"Cost of product revenue",cogs_prod,ind=1); r+=1
    _fs(ws,r,"Cost of service revenue",cogs_svc,ind=1,ul=True); r+=1
    _fs(ws,r,"Gross profit",gp,bold=True); r+=2
    _fs(ws,r,"Operating expenses",hdr=True); r+=1
    _fs(ws,r,"Research and development",rd,ind=1); r+=1
    _fs(ws,r,"Sales and marketing",sm,ind=1); r+=1
    _fs(ws,r,"General and administrative",ga,ind=1,ul=True); r+=1
    _fs(ws,r,"Total operating expenses",sga,bold=True); r+=2
    _fs(ws,r,"Earnings before interest, taxes, depreciation and amortization (EBITDA)",E,bold=True); r+=1
    _fs(ws,r,"Depreciation and amortization",da,ind=1,ul=True); r+=1
    _fs(ws,r,"Operating income",ebit,bold=True); r+=2
    _fs(ws,r,"Other income (expense)",hdr=True); r+=1
    _fs(ws,r,"Interest income",int_inc,ind=1); r+=1
    _fs(ws,r,"Interest expense",-int_exp,ind=1,ul=True); r+=1
    _fs(ws,r,"Total other income (expense), net",-interest_net,bold=True); r+=2
    _fs(ws,r,"Income before provision for income taxes",ebt,bold=True); r+=1
    _fs(ws,r,"Provision for income taxes (effective rate ~30%)",-tax,ind=1,ul=True); r+=1
    _fs(ws,r,"Net income",N,bold=True); r+=2
    _fs(ws,r,f"Earnings per share — Basic ${N/S:.2f}  |  Diluted ${N/(S*1.06):.2f}  |  Wtd-avg shares: Basic {S:.0f}M, Diluted {S*1.06:.1f}M",note=True); r+=3
    # ── BS ──────────────────────────────────────────────────────────
    _fs_title(ws,r,"CONSOLIDATED BALANCE SHEET"); r+=1
    _fs_title(ws,r,"(Amounts in millions of U.S. dollars — December 31, 2024)",sub=True); r+=2
    _fs(ws,r,"ASSETS",hdr=True); r+=1
    _fs(ws,r,"Current assets",bold=True,ind=0); r+=1
    _fs(ws,r,"Cash and cash equivalents",cash_cce,ind=1); r+=1
    _fs(ws,r,"Short-term investments",cash_stinv,ind=1); r+=1
    _fs(ws,r,"Total cash, cash equivalents and short-term investments",cash_tot,ind=2,bold=True); r+=1
    _fs(ws,r,"Accounts receivable, net of allowances of $2.1",ar,ind=1); r+=1
    _fs(ws,r,"Inventories",inv,ind=1); r+=1
    _fs(ws,r,"Prepaid expenses and other current assets",prepaid,ind=1); r+=1
    _fs(ws,r,"Income taxes receivable",tax_rec,ind=1,ul=True); r+=1
    _fs(ws,r,"Total current assets",curr_a,bold=True); r+=2
    _fs(ws,r,"Property and equipment, net",ppe); r+=1
    _fs(ws,r,"Goodwill",goodwill); r+=1
    _fs(ws,r,"Intangible assets, net",intang); r+=1
    _fs(ws,r,"Other long-term assets",other_lta,ul=True); r+=1
    _fs(ws,r,"Total assets",tot_a,bold=True); r+=3
    _fs(ws,r,"LIABILITIES AND STOCKHOLDERS' EQUITY",hdr=True); r+=1
    _fs(ws,r,"Current liabilities",bold=True); r+=1
    _fs(ws,r,"Accounts payable",ap,ind=1); r+=1
    _fs(ws,r,"Accrued compensation and benefits",accrued,ind=1); r+=1
    _fs(ws,r,"Other accrued liabilities",oth_acc,ind=1); r+=1
    _fs(ws,r,"Current portion of long-term debt",std,ind=1); r+=1
    _fs(ws,r,"Income taxes payable",tax_pay,ind=1,ul=True); r+=1
    _fs(ws,r,"Total current liabilities",curr_l,bold=True); r+=2
    _fs(ws,r,"Long-term debt, net of current portion",ltd); r+=1
    _fs(ws,r,"Deferred revenue, non-current",def_rev); r+=1
    _fs(ws,r,"Other long-term liabilities",oth_ltl,ul=True); r+=1
    _fs(ws,r,"Total liabilities",tot_l,bold=True); r+=2
    _fs(ws,r,"Stockholders' equity",bold=True); r+=1
    _fs(ws,r,"Common stock and additional paid-in capital",cs,ind=1); r+=1
    _fs(ws,r,"Retained earnings",re,ind=1); r+=1
    _fs(ws,r,"Accumulated other comprehensive income (loss)",aoci,ind=1,ul=True); r+=1
    _fs(ws,r,"Total stockholders' equity",tot_eq,bold=True); r+=1
    _fs(ws,r,"Total liabilities and stockholders' equity",tot_a,bold=True); r+=3
    # ── CFS ─────────────────────────────────────────────────────────
    _fs_title(ws,r,"CONSOLIDATED STATEMENT OF CASH FLOWS"); r+=1
    _fs_title(ws,r,"(Amounts in millions of U.S. dollars — Fiscal Year 2024)",sub=True); r+=2
    _fs(ws,r,"Cash flows from operating activities",hdr=True); r+=1
    _fs(ws,r,"Net income",N,ind=1); r+=1
    _fs(ws,r,"Depreciation and amortization",da,ind=2); r+=1
    _fs(ws,r,"Stock-based compensation expense",sbc,ind=2); r+=1
    _fs(ws,r,"Amortization of deferred financing costs",round(R*.002,1),ind=2); r+=1
    _fs(ws,r,"Accounts receivable",d_ar,ind=2); r+=1
    _fs(ws,r,"Inventories",d_inv,ind=2); r+=1
    _fs(ws,r,"Accounts payable",d_ap,ind=2); r+=1
    _fs(ws,r,"Accrued liabilities and other",oth_wc,ind=2,ul=True); r+=1
    _fs(ws,r,"Net cash provided by operating activities",cfo,bold=True); r+=2
    _fs(ws,r,"Cash flows from investing activities",hdr=True); r+=1
    _fs(ws,r,"Capital expenditures",-capex,ind=1); r+=1
    _fs(ws,r,"Purchases of short-term investments",round(-R*.04,1),ind=1); r+=1
    _fs(ws,r,"Proceeds from maturities of investments",round(R*.016,1),ind=1,ul=True); r+=1
    _fs(ws,r,"Net cash used in investing activities",cfi,bold=True); r+=2
    _fs(ws,r,"Cash flows from financing activities",hdr=True); r+=1
    _fs(ws,r,"Proceeds from exercise of stock options",sop,ind=1); r+=1
    _fs(ws,r,"Taxes paid related to vested stock awards",tax_sop,ind=1); r+=1
    _fs(ws,r,"Repayment of long-term debt",debt_pay,ind=1); r+=1
    _fs(ws,r,"Dividends paid to stockholders",-divs,ind=1,ul=True); r+=1
    _fs(ws,r,"Net cash used in financing activities",cff,bold=True); r+=2
    _fs(ws,r,"Net increase (decrease) in cash and cash equivalents",net_chg,bold=True); r+=1
    _fs(ws,r,"Cash and cash equivalents, beginning of year",beg_cash); r+=1
    _fs(ws,r,"Cash and cash equivalents, end of year",cash_cce,bold=True,ul=True); r+=2
    _fs(ws,r,f"Supplemental: Capital expenditures above = cash paid for PP&E. Dividends per share: ${divs/S:.2f}.",note=True)


def _make_adv_dcf(wb, co):
    R=co["rev_n"]; E=co["ebitda_n"]; N=co["ni_n"]; S=co["shares_n"]; nm=co["name"]
    P=co["price_n"]; g=co.get("rev_growth",.05); wc_r=co.get("wacc",.0935)
    tg=co.get("terminal_g",.025); TAX=0.30
    cogs=round(R*.60,1); sga=round(R*.40-E,1); da=round(R*.06,1)
    ebit=E-da; ebt=round(N/(1-TAX),1); interest_net=round(ebit-ebt,1); tax=round(ebt*TAX,1)
    nd=round(E*1.05/5)*5; mktcap=round(P*S,1); ev=mktcap+nd
    ltd_sr=round(nd*.72,1); ltd_notes=round(nd*.28,1)
    kd_pre=round(.07,4); ke=round(.045+1.10*.055,4)
    prod_rev=round(R*.70,1); svc_rev=round(R-prod_rev,1)
    cogs_prod=round(cogs*.80,1); cogs_svc=round(cogs-cogs_prod,1)
    rd=round(sga*.40,1); sm=round(sga*.35,1); ga=round(sga-rd-sm,1)
    int_inc=round(ebt*.02,1); int_exp=round(interest_net+int_inc,1)

    ws=wb.create_sheet("Advanced Practice"); _cw(ws,[("A",50),("B",14)])
    r=1; _adv_header(ws,r,nm); r+=2
    # IS
    _fs_title(ws,r,nm); r+=1
    _fs_title(ws,r,"CONSOLIDATED STATEMENT OF OPERATIONS  (USD millions — FY 2024)",sub=True); r+=2
    _fs(ws,r,"Revenues",hdr=True); r+=1
    _fs(ws,r,"Product revenue",prod_rev,ind=1); r+=1
    _fs(ws,r,"Service and subscription revenue",svc_rev,ind=1,ul=True); r+=1
    _fs(ws,r,"Total revenues",R,bold=True); r+=2
    _fs(ws,r,"Cost of product revenue",cogs_prod,ind=1); r+=1
    _fs(ws,r,"Cost of service revenue",cogs_svc,ind=1,ul=True); r+=1
    _fs(ws,r,"Gross profit",round(R-cogs,1),bold=True); r+=2
    _fs(ws,r,"Research and development",rd,ind=1); r+=1
    _fs(ws,r,"Sales and marketing",sm,ind=1); r+=1
    _fs(ws,r,"General and administrative",ga,ind=1,ul=True); r+=1
    _fs(ws,r,"Total operating expenses (excl. D&A)",sga,bold=True); r+=1
    _fs(ws,r,"EBITDA",E,bold=True); r+=1
    _fs(ws,r,"Depreciation and amortization",da,ind=1,ul=True); r+=1
    _fs(ws,r,"Operating income (EBIT)",ebit,bold=True); r+=1
    _fs(ws,r,"Interest income",int_inc,ind=1); r+=1
    _fs(ws,r,"Interest expense",-int_exp,ind=1,ul=True); r+=1
    _fs(ws,r,"Income before taxes",ebt,bold=True); r+=1
    _fs(ws,r,"Provision for income taxes",-tax,ind=1,ul=True); r+=1
    _fs(ws,r,"Net income",N,bold=True); r+=2
    _fs(ws,r,f"Diluted EPS: ${N/(S*1.06):.2f}  |  Diluted shares: {S*1.06:.1f}M  |  Basic EPS: ${N/S:.2f}",note=True); r+=3
    # Market & Capital Structure Data
    _fs_title(ws,r,"SELECTED MARKET & CAPITAL STRUCTURE DATA"); r+=1
    _fs_title(ws,r,"(As of December 31, 2024)",sub=True); r+=2
    _fs(ws,r,"Market Data",hdr=True); r+=1
    _fs(ws,r,"Share price (closing)",P,ind=1); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=1
    _fs(ws,r,"Diluted shares outstanding (M)",S*1.06,ind=1); r+=1
    _fs(ws,r,"Market capitalization ($M)",mktcap,ind=1,bold=True); r+=1
    _fs(ws,r,f"52-week trading range:  ${round(P*.72,2):.2f}  —  ${round(P*1.14,2):.2f}",note=True,ind=1); r+=2
    _fs(ws,r,"Capital Structure",hdr=True); r+=1
    _fs(ws,r,"Cash and short-term investments ($M)",round(R*.15,1),ind=1); r+=1
    _fs(ws,r,"Total debt ($M)",nd+round(R*.15,1),ind=1); r+=1
    _fs(ws,r,"Net debt ($M)",nd,ind=1,bold=True); r+=1
    _fs(ws,r,"Enterprise value ($M)",ev,ind=1,bold=True); r+=2
    _fs(ws,r,"Debt Structure",hdr=True); r+=1
    _fs(ws,r,f"Senior secured term loan due 2028 — ${ltd_sr}M  @  SOFR + 2.50% (all-in ~6.5%)",note=False,ind=1); r+=1
    _fs(ws,r,f"Senior unsecured notes due 2031 — ${ltd_notes}M  @  6.25% fixed",ind=1); r+=1
    _fs(ws,r,"Pre-tax weighted-avg. cost of debt",kd_pre,ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=1
    _fs(ws,r,"After-tax cost of debt (×0.70)",round(kd_pre*.70,4),ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=2
    _fs(ws,r,"WACC Inputs from Equity Research Coverage",hdr=True); r+=1
    _fs(ws,r,"10-Year U.S. Treasury yield (risk-free rate)",.045,ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=1
    _fs(ws,r,"Equity risk premium (ERP)",.055,ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=1
    _fs(ws,r,"Company levered beta (5-year monthly regression)",1.10,ind=1); r+=1
    _fs(ws,r,"Implied cost of equity  [= Rf + β × ERP]",ke,ind=1,bold=True); ws[f"B{r-1}"].number_format="0.00%"; r+=1
    _fs(ws,r,"Debt / (Debt + Equity) at market value",round(nd/ev,4),ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=1
    _fs(ws,r,"Equity / (Debt + Equity) at market value",round(mktcap/ev,4),ind=1); ws[f"B{r-1}"].number_format="0.00%"; r+=2
    _fs(ws,r,f"Analyst consensus revenue growth (next 3 years avg.): {g*100:.1f}%/yr  |  Long-run GDP growth consensus: {tg*100:.1f}%",note=True); r+=1
    _fs(ws,r,"Terminal growth rate should reflect long-run nominal GDP and MUST be less than WACC or model produces infinite value.",note=True)


def _make_adv_comps(wb, co):
    E=co["ebitda_n"]; S=co["shares_n"]; P=co["price_n"]; nm=co["name"]
    nd=round(E*1.05/5)*5; mktcap=round(P*S,1)
    # Peer data: (name, mktcap, net_debt, ebitda, revenue, net_income, rev_growth, ebitda_margin)
    peers=[
        ("TechAlpha Corp",    1202, 1104-1202,   120, 480, 65,  .08, .250),
        ("DataCore Systems",   810,  810-810,    100, 420, 50,  .05, .238),
        ("CloudSystems Inc",  1239, 1260-1239,   120, 400, 59,  .12, .300),
        ("InfoPro Ltd",        853,  858-853,    110, 510, 54,  .04, .216),
    ]
    ws=wb.create_sheet("Advanced Practice")
    _cw(ws,[("A",22),("B",12),("C",12),("D",12),("E",12),("F",12),("G",10),("H",10),("I",10),("J",10)])
    r=1; _adv_header(ws,r,nm); r+=2
    _fs_title(ws,r,"COMPARABLE COMPANY TRADING ANALYSIS — EQUITY RESEARCH SCREEN"); r+=1
    _fs_title(ws,r,"(USD millions except per share; market data as of December 31, 2024)",sub=True); r+=2
    # Header row
    hdrs=["Company","Mkt Cap","Net Debt","Total Debt","LTM Revenue","LTM EBITDA","LTM Net Inc.","Rev Growth","EBITDA Mgn","Share Price"]
    for i,h in enumerate(hdrs,1):
        c=ws.cell(row=r,column=i,value=h)
        c.font=Font(bold=True,size=8,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1E3A5F")
        c.alignment=Alignment(horizontal="center",wrap_text=True)
    ws.row_dimensions[r].height=28; r+=1
    # Peer rows
    peer_prices=[round(mc/s,2) for mc,s in [(1202,26.7),(810,28.9),(1239,21.0),(853,15.8)]]
    peer_shares=[26.7,28.9,21.0,15.8]
    for i,((pnm,mc,net_d,ebitda,rev,ni,rg,em),pp,ps) in enumerate(zip(peers,peer_prices,peer_shares)):
        td=round(mc+net_d,0)
        cols=[pnm,mc,net_d,td,rev,ebitda,ni,f"{rg*100:.1f}%",f"{em*100:.1f}%",f"${pp:.2f}"]
        for j,val in enumerate(cols,1):
            c=ws.cell(row=r,column=j,value=val)
            c.font=Font(size=9); c.alignment=Alignment(horizontal="right" if j>1 else "left")
            if j>1 and isinstance(val,(int,float)):
                c.number_format='#,##0.0'
        r+=1
    r+=1
    # Subject company data (noisier — more line items than needed)
    _fs_title(ws,r,f"SUBJECT COMPANY: {nm}"); r+=1
    _fs_title(ws,r,"Selected Financial & Market Data  (USD millions — FY 2024)",sub=True); r+=2
    _fs(ws,r,"Market capitalization",mktcap,bold=True); r+=1
    _fs(ws,r,"Net debt (total debt minus cash & equivalents)",nd); r+=1
    _fs(ws,r,"Enterprise value  [market cap + net debt]",mktcap+nd,bold=True); r+=2
    _fs(ws,r,"LTM Total revenues",co["rev_n"]); r+=1
    _fs(ws,r,"LTM EBITDA",E,bold=True); r+=1
    _fs(ws,r,"LTM Operating income (EBIT)",round(E-co["rev_n"]*.06,1)); r+=1
    _fs(ws,r,"LTM Net income",co["ni_n"]); r+=1
    _fs(ws,r,"LTM Diluted EPS",round(co["ni_n"]/(S*1.06),2)); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=2
    _fs(ws,r,"Shares outstanding — Basic (M)",S); r+=1
    _fs(ws,r,"Shares outstanding — Diluted (M)",round(S*1.06,1)); r+=1
    _fs(ws,r,"Share price",P); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=2
    _fs(ws,r,"NOTE: Derive EV/EBITDA and EV/Revenue multiples for each peer. Apply peer median multiples to subject company to arrive at implied Enterprise Value and implied share price.",note=True)


def _make_adv_precedent(wb, co):
    E=co["ebitda_n"]; S=co["shares_n"]; P=co["price_n"]; nm=co["name"]
    nd=round(E*1.05/5)*5
    ws=wb.create_sheet("Advanced Practice"); _cw(ws,[("A",50),("B",14)])
    r=1; _adv_header(ws,r,nm); r+=2
    _fs_title(ws,r,"PRECEDENT TRANSACTION RESEARCH — M&A DEAL SUMMARIES"); r+=1
    _fs_title(ws,r,"Selected completed acquisitions in the sector  (USD millions unless noted)",sub=True); r+=3

    deals=[
        ("AcquireCo / TechAlpha Corporation","February 14, 2022",1320,120,975,
         "Strategic acquisition to expand cloud infrastructure capabilities. AcquireCo cited ~$45M in annual run-rate synergies (cost savings + cross-sell). "
         "Premium of 35.4% to TechAlpha's unaffected 30-day VWAP of $36.50. Deal financed with $800M of new term loans and $520M equity."),
        ("MegaCorp Industries / DataCore Systems","September 8, 2021",980,100,809,
         "Bolt-on acquisition. MegaCorp CEO: 'Accelerates our data analytics roadmap by 3 years.' "
         "Unaffected market cap $809M; deal at 21.1% premium. All-cash offer at $33.90 per DataCore share. LTM EBITDA at signing: $100M; no NTM guidance available."),
        ("GlobalTech Partners / InfoPro Ltd","June 3, 2023",1045,110,856,
         "Public-to-private buyout led by GlobalTech with co-investment from two PE firms. "
         "InfoPro had been exploring strategic alternatives since Q4 2022. Implied EV/LTM EBITDA premium vs. InfoPro's 90-day avg. trading multiple: +2.8 turns. "
         "Deal included $180M rollover equity from InfoPro management."),
    ]
    for deal_nm,date,dev,debitda,predeal_mc,desc in deals:
        _fs(ws,r,f"DEAL:  {deal_nm}",hdr=True); r+=1
        _fs(ws,r,f"Announcement Date: {date}",note=True,ind=1); r+=1
        _fs(ws,r,"Deal Enterprise Value (EV)",dev,ind=1,bold=True); r+=1
        _fs(ws,r,"Target LTM EBITDA at announcement",debitda,ind=1); r+=1
        _fs(ws,r,"Target unaffected market capitalization",predeal_mc,ind=1); r+=1
        _fs(ws,r,"Total equity consideration",round(predeal_mc*1.354 if deal_nm.startswith("Acq") else (predeal_mc*1.211 if deal_nm.startswith("Mega") else predeal_mc*1.220),1),ind=1); r+=1
        for line in [desc[i:i+110] for i in range(0,len(desc),110)]:
            _fs(ws,r,line,note=True,ind=1); r+=1
        r+=2

    _fs_title(ws,r,f"SUBJECT COMPANY: {nm}"); r+=1
    _fs_title(ws,r,"Current trading and financial profile  (USD millions — FY 2024)",sub=True); r+=2
    _fs(ws,r,"LTM EBITDA",E,bold=True); r+=1
    _fs(ws,r,"Current share price",P); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=1
    _fs(ws,r,"Shares outstanding (M)",S); r+=1
    _fs(ws,r,"Current market capitalization",round(P*S,1),bold=True); r+=1
    _fs(ws,r,"Net debt",nd); r+=1
    _fs(ws,r,"Current Enterprise Value",round(P*S+nd,1),bold=True); r+=2
    _fs(ws,r,"NOTE: Compute EV/EBITDA for each deal. Apply median deal multiple to subject EBITDA to get implied acquisition EV → subtract net debt → divide by shares → floor acquisition price per share.",note=True)


def _make_adv_lbo(wb, co):
    R=co["rev_n"]; E=co["ebitda_n"]; N=co["ni_n"]; nm=co["name"]
    da=round(R*.06,1); capex=round(R*.07,1); g=co.get("rev_growth",.05); TAX=0.30
    ebit=E-da; ebt=round(N/(1-TAX),1); tax=round(ebt*TAX,1)
    cogs=round(R*.60,1); sga=round(R*.40-E,1)
    entry_ev=round(E*9.0,1); debt=round(entry_ev*.60,1); equity_in=round(entry_ev*.40,1)
    interest=round(debt*.07,1)

    ws=wb.create_sheet("Advanced Practice"); _cw(ws,[("A",50),("B",14)])
    r=1; _adv_header(ws,r,nm); r+=2
    _fs_title(ws,r,f"LEVERAGED BUYOUT — {nm.upper()}"); r+=1
    _fs_title(ws,r,"Management Presentation & Lender Materials  (USD millions — FY 2024)",sub=True); r+=3

    _fs_title(ws,r,"SECTION 1:  TARGET COMPANY FINANCIALS"); r+=2
    _fs(ws,r,"Income Statement Summary",hdr=True); r+=1
    _fs(ws,r,"Total revenues",R,bold=True); r+=1
    _fs(ws,r,"Cost of revenues",cogs,ind=1); r+=1
    _fs(ws,r,"Gross profit",round(R-cogs,1)); r+=1
    _fs(ws,r,"Operating expenses (R&D, S&M, G&A)",sga,ind=1); r+=1
    _fs(ws,r,"EBITDA",E,bold=True); r+=1
    _fs(ws,r,"D&A",da,ind=1); r+=1
    _fs(ws,r,"EBIT",ebit); r+=1
    _fs(ws,r,"Interest expense (pre-transaction)",round(ebt*.20,1),ind=1); r+=1
    _fs(ws,r,"Pre-tax income",ebt); r+=1
    _fs(ws,r,"Income taxes",tax,ind=1); r+=1
    _fs(ws,r,"Net income",N,bold=True); r+=2
    _fs(ws,r,"Key Metrics",hdr=True); r+=1
    _fs(ws,r,"EBITDA margin",round(E/R*100,1),ind=1); ws[f"B{r-1}"].number_format='0.0"%"'; r+=1
    _fs(ws,r,"Capex ($M)",capex,ind=1); r+=1
    _fs(ws,r,"Historical revenue growth (3-yr CAGR)",round(g*100,1),ind=1); ws[f"B{r-1}"].number_format='0.0"%"'; r+=1
    _fs(ws,r,f"Management projects {g*100:.0f}–{g*100+2:.0f}%/yr revenue growth over next 5 years; EBITDA margins stable.",note=True,ind=1); r+=3

    _fs_title(ws,r,"SECTION 2:  TRANSACTION & FINANCING TERMS"); r+=2
    _fs(ws,r,"Transaction Structure",hdr=True); r+=1
    _fs(ws,r,"Acquisition structure",ind=1); ws[f"A{r-1}"].value="    All-equity public-to-private; no existing debt assumed"; r+=1
    _fs(ws,r,"Entry valuation multiple (EV / LTM EBITDA)",9.0,ind=1); r+=1
    _fs(ws,r,"Implied Entry Enterprise Value",entry_ev,ind=1,bold=True); r+=1
    _fs(ws,r,"Transaction fees and expenses (est.)",round(entry_ev*.02,1),ind=1); r+=2
    _fs(ws,r,"Financing Structure (Sources & Uses)",hdr=True); r+=1
    _fs(ws,r,"Senior Secured Term Loan B  —  60% of EV  @  SOFR+350bps (~7.0% all-in)",debt,ind=1); r+=1
    _fs(ws,r,"Sponsor equity contribution  —  40% of EV",equity_in,ind=1,bold=True); r+=2
    _fs(ws,r,"Debt Terms",hdr=True); r+=1
    _fs(ws,r,"Amortization","1% per annum of original principal (bullet at maturity)",ind=1); ws[f"B{r-1}"].value="1% p.a. + cash sweep"; r+=1
    _fs(ws,r,"Cash sweep: 75% of excess cash flow applied to debt paydown",note=True,ind=1); r+=1
    _fs(ws,r,"Maturity","7 years",ind=1); ws[f"B{r-1}"].value="7 years from close"; r+=1
    _fs(ws,r,"Financial covenant: max Net Debt / EBITDA of 6.5×, stepping down 0.5× per year",note=True,ind=1); r+=2
    _fs(ws,r,"Exit Assumptions",hdr=True); r+=1
    _fs(ws,r,"Assumed hold period","5 years",ind=1); ws[f"B{r-1}"].value="5 years"; r+=1
    _fs(ws,r,"Exit multiple (conservative base case: entry = exit)",9.0,ind=1); r+=1
    _fs(ws,r,"Target IRR for sponsor capital: 20%+ (minimum acceptable)",note=True,ind=1); r+=2
    _fs(ws,r,"NOTE: Build Sources & Uses → Debt Schedule (5 yrs of beginning debt, interest, EBITDA, FCF for paydown, ending debt) → Exit & Returns (Exit EV, remaining debt, exit equity, MOIC, IRR).",note=True)


def _make_adv_merger(wb, co):
    R=co["rev_n"]; E=co["ebitda_n"]; N=co["ni_n"]; P=co["price_n"]; S=co["shares_n"]; nm=co["name"]
    TAX=0.30; cogs=round(R*.60,1); sga=round(R*.40-E,1); da=round(R*.06,1)
    ebit=E-da; ebt=round(N/(1-TAX),1); tax=round(ebt*TAX,1)
    tgt_rev=round(R*.60,1); tgt_ebitda=round(E*.58,1); tgt_ni=round(N*.35,1)
    tgt_P=round(P*.60,2); tgt_S=round(S*.40,1); tgt_da=round(R*.60*.06,1)
    tgt_cogs=round(tgt_rev*.60,1); tgt_sga=round(tgt_rev*.40-tgt_ebitda,1)
    tgt_ebit=tgt_ebitda-tgt_da; tgt_ebt=round(tgt_ni/(1-TAX),1)
    tgt_int=round(tgt_ebit-tgt_ebt,1); tgt_tax=round(tgt_ebt*TAX,1)
    syn=round(N*.08,1); amort=round(N*.04,1)

    ws=wb.create_sheet("Advanced Practice"); _cw(ws,[("A",50),("B",14),("C",14)])
    r=1; _adv_header(ws,r,nm); r+=2
    _fs_title(ws,r,f"MERGER ANALYSIS — {nm.upper()} (ACQUIRER) + TARGETCO (TARGET)"); r+=1
    _fs_title(ws,r,"Pro Forma Merger Analysis  (USD millions, FY 2024 financials)",sub=True); r+=3

    # Side-by-side IS
    _fs_title(ws,r,"INCOME STATEMENTS — STANDALONE"); r+=1
    ws.cell(row=r,column=2,value=nm).font=Font(bold=True,size=9)
    ws.cell(row=r,column=3,value="TargetCo").font=Font(bold=True,size=9); r+=1
    rows_is=[("Total revenues",R,tgt_rev),("Cost of revenues",cogs,tgt_cogs),
             ("Gross profit",round(R-cogs,1),round(tgt_rev-tgt_cogs,1)),
             ("Operating expenses (excl. D&A)",sga,tgt_sga),("EBITDA",E,tgt_ebitda),
             ("Depreciation & amortization",da,tgt_da),("EBIT",ebit,tgt_ebit),
             ("Interest expense (net)",round(ebt*.20,1),tgt_int),
             ("Income before taxes",ebt,tgt_ebt),("Income taxes",tax,tgt_tax),
             ("Net income",N,tgt_ni)]
    for lbl,va,vb in rows_is:
        bold=lbl in ("Total revenues","Gross profit","EBITDA","Net income")
        ws.cell(row=r,column=1,value=lbl).font=Font(bold=bold,size=9)
        for j,v in [(2,va),(3,vb)]:
            c=ws.cell(row=r,column=j,value=v)
            c.font=Font(bold=bold,size=9); c.number_format='#,##0.0'; c.alignment=_RIGHT_ALIGN
        r+=1
    r+=2

    _fs_title(ws,r,"SHARE & DEAL DATA"); r+=2
    _fs(ws,r,f"{nm} — Share Data",hdr=True); r+=1
    _fs(ws,r,"Share price",P,ind=1); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=1
    _fs(ws,r,"Basic shares outstanding (M)",S,ind=1); r+=1
    _fs(ws,r,"Market capitalization ($M)",round(P*S,1),ind=1,bold=True); r+=2
    _fs(ws,r,"TargetCo — Share & Deal Data",hdr=True); r+=1
    _fs(ws,r,"Unaffected market price (30-day VWAP pre-announcement)",tgt_P,ind=1); ws[f"B{r-1}"].number_format='"$"#,##0.00'; r+=1
    _fs(ws,r,"Shares outstanding (M)",tgt_S,ind=1); r+=1
    _fs(ws,r,"Unaffected market capitalization ($M)",round(tgt_P*tgt_S,1),ind=1); r+=1
    _fs(ws,r,"Proposed acquisition premium",0.25,ind=1); ws[f"B{r-1}"].number_format="0%"; r+=1
    _fs(ws,r,"Consideration: all-stock (new shares of Acquirer issued to Target shareholders)",ind=1); ws[f"B{r-1}"].value="All-stock"; r+=2
    _fs(ws,r,"Deal Synergies & Purchase Accounting",hdr=True); r+=1
    _fs(ws,r,"Estimated after-tax run-rate cost synergies (Year 2 onwards)",syn,ind=1,bold=True); r+=1
    _fs(ws,r,"  — Sources: procurement savings $X, duplicate headcount $Y, facility consolidation $Z",note=True,ind=1); r+=1
    _fs(ws,r,"Intangible assets acquired (customer relationships, technology, brand)",round(amort*25,1),ind=1); r+=1
    _fs(ws,r,"Intangible amortization (straight-line, 25-year life, non-cash P&L charge)",-amort,ind=1,bold=True); r+=2
    _fs(ws,r,"NOTE: Compute offer price per share, total deal value, new shares issued, pro forma NI (acquirer + target + synergies − intangibles amort), pro forma shares, and test EPS accretion/dilution.",note=True)


def _make_adv_budget(wb, co):
    R=co["rev_n"]; E=co["ebitda_n"]; nm=co["name"]
    cogs=round(R*.60,1); sga=round(R*.40-E,1); TAX=0.30
    # Budget vs actual with MORE detail than model needs (by product line)
    UB,UA=12_000,11_500; rb=round(R*.24,1); ra=round(rb*.942,1)
    cb=round(cogs*.24,1); ca=round(cb*.976,1)
    sb=round(sga*.25,1); sa=round(sb*1.075,1)
    pb=round(rb*1e6/UB); pa=round(ra*1e6/UA)
    # Sub-line detail (noise)
    rb_prod=round(rb*.72,1); rb_svc=round(rb-rb_prod,1)
    ra_prod=round(ra*.73,1); ra_svc=round(ra-ra_prod,1)
    cb_prod=round(cb*.80,1); cb_svc=round(cb-cb_prod,1)
    ca_prod=round(ca*.79,1); ca_svc=round(ca-ca_prod,1)
    sb_rd=round(sb*.40,1); sb_sm=round(sb*.35,1); sb_ga=round(sb-sb_rd-sb_sm,1)
    sa_rd=round(sa*.42,1); sa_sm=round(sa*.36,1); sa_ga=round(sa-sa_rd-sa_sm,1)
    # derived
    rb_gp=round(rb-cb,1); ra_gp=round(ra-ca,1)
    rb_eb=round(rb_gp-sb,1); ra_eb=round(ra_gp-sa,1)

    ws=wb.create_sheet("Advanced Practice")
    _cw(ws,[("A",32),("B",12),("C",12),("D",12),("E",12),("F",10)])
    r=1; _adv_header(ws,r,nm); r+=2
    _fs_title(ws,r,f"{nm}  —  Q1 MANAGEMENT PROFIT & LOSS REPORT"); r+=1
    _fs_title(ws,r,"Internal Use Only  |  All figures in $M  |  Unit economics in actual $",sub=True); r+=2

    # Table header
    hdrs=["","Budget","Actual","Variance $","Variance %","F / U"]
    for j,h in enumerate(hdrs,1):
        c=ws.cell(row=r,column=j,value=h)
        c.font=Font(bold=True,size=9,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1E3A5F")
        c.alignment=Alignment(horizontal="center" if j>1 else "left")
    r+=1

    def _bv(ws,row,lbl,bv,av,bold=False,ind=0):
        ws.cell(row=row,column=1,value="  "*ind+lbl).font=Font(bold=bold,size=9)
        for j,v in [(2,bv),(3,av)]:
            c=ws.cell(row=row,column=j,value=v)
            c.font=Font(bold=bold,size=9); c.number_format='#,##0.0'; c.alignment=_RIGHT_ALIGN
        # Leave variance cols blank for student to compute

    _bv(ws,r,"REVENUES",None,None,bold=True); r+=1
    _bv(ws,r,"Product revenue",rb_prod,ra_prod,ind=1); r+=1
    _bv(ws,r,"Service revenue",rb_svc,ra_svc,ind=1); r+=1
    _bv(ws,r,"Total revenues",rb,ra,bold=True); r+=2
    _bv(ws,r,"COST OF REVENUES",None,None,bold=True); r+=1
    _bv(ws,r,"Cost of product revenue",cb_prod,ca_prod,ind=1); r+=1
    _bv(ws,r,"Cost of service revenue",cb_svc,ca_svc,ind=1); r+=1
    _bv(ws,r,"Total cost of revenues",cb,ca,bold=True); r+=1
    _bv(ws,r,"Gross profit",rb_gp,ra_gp,bold=True); r+=2
    _bv(ws,r,"OPERATING EXPENSES",None,None,bold=True); r+=1
    _bv(ws,r,"Research & development",sb_rd,sa_rd,ind=1); r+=1
    _bv(ws,r,"Sales & marketing",sb_sm,sa_sm,ind=1); r+=1
    _bv(ws,r,"General & administrative",sb_ga,sa_ga,ind=1); r+=1
    _bv(ws,r,"Total operating expenses",sb,sa,bold=True); r+=1
    _bv(ws,r,"EBITDA",rb_eb,ra_eb,bold=True); r+=3

    _fs_title(ws,r,"UNIT ECONOMICS  —  Revenue Bridge"); r+=2
    _bv(ws,r,"Units sold",UB,UA); r+=1
    _bv(ws,r,"Avg. selling price / unit ($)",pb,pa); r+=2
    _fs(ws,r,"NOTE: Build the variance table above (Variance $, Variance %, F/U flag for each line). "
           "Then decompose the revenue miss into Volume Effect and Price Effect. "
           "Aggregate sub-line items to match a simplified model (Revenue, COGS, Gross Profit, SG&A, EBITDA).",note=True)


_ADV_BUILDERS = {
    "3stmt": _make_adv_3stmt, "dcf": _make_adv_dcf, "comps": _make_adv_comps,
    "precedent": _make_adv_precedent, "lbo": _make_adv_lbo,
    "merger": _make_adv_merger, "budget": _make_adv_budget,
}


def make_model_excel_advanced(model_key, co):
    """Complete Model (reference) + Advanced Practice (given inputs only, blank workspace)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Add complete reference sheets
    {"3stmt": _make_3stmt_excel, "dcf": _make_dcf_excel, "comps": _make_comps_excel,
     "precedent": _make_precedent_excel, "lbo": _make_lbo_excel,
     "merger": _make_merger_excel, "budget": _make_budget_excel
    }[model_key](wb, co)
    # Remove the guided Practice sheets (keep only Complete sheets)
    for sname in [s for s in wb.sheetnames if "Practice" in s]:
        wb.remove(wb[sname])
    # Add the blank Advanced Practice sheet
    _ADV_BUILDERS[model_key](wb, co)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def _build_tab_header(model_key, sheet_desc):
    co = current_company()
    col_a, col_b = st.columns(2)
    with col_a:
        buf = make_model_excel(model_key, co)
        st.download_button(
            f"⬇  Download Excel Template  —  {co['name']}",
            buf,
            file_name=f"{model_key}_{co['name'].replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"dl_guided_{model_key}",
        )
    with col_b:
        adv_buf = make_model_excel_advanced(model_key, co)
        st.download_button(
            f"🎯  Advanced Modeling  —  {co['name']}",
            adv_buf,
            file_name=f"{model_key}_{co['name'].replace(' ','_')}_advanced.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"dl_adv_{model_key}",
        )
    st.markdown(
        f'<div style="font-size:.78rem;color:#4B5563;padding:5px 0 10px;">'
        f'<b>Excel Template:</b> {sheet_desc} &nbsp;·&nbsp; '
        f'<b style="color:#1D4ED8;">■</b> Blue = given input &nbsp;·&nbsp; '
        f'<span style="background:#FFFBEB;padding:0 4px;border:1px solid #FDE68A;">Yellow</span>'
        f' = write your formula &nbsp;|&nbsp; '
        f'<b>Advanced:</b> Complete reference + blank workspace — no structure provided, start from scratch.</div>',
        unsafe_allow_html=True)
    st.markdown("---")


# ── Module header ──────────────────────────────────────────────────────────────
_PAGE_META = {
    "overview":  ("Financial Modeling School",       "Seven core Excel models — IB & PE curriculum"),
    "3stmt":     ("① Three Statement Model",          "Foundation · Income Statement · Balance Sheet · Cash Flow"),
    "dcf":       ("② DCF Model",                     "Valuation · Discounted Cash Flow · WACC · Terminal Value"),
    "comps":     ("③ Comparable Company Analysis",    "Valuation · Public Market Multiples · Peer Benchmarking"),
    "precedent": ("④ Precedent Transactions",         "Valuation · M&A Deal Multiples · Control Premium"),
    "lbo":       ("⑤ LBO Model",                     "Advanced · Sources & Uses · Debt Schedule · MOIC / IRR"),
    "merger":    ("⑥ Merger Model",                  "Advanced · Accretion / Dilution · EPS Impact"),
    "budget":    ("⑦ Budget vs. Actual",              "FP&A · Variance Analysis · Volume / Price Decomposition"),
}

def page_header(page):
    title, subtitle = _PAGE_META.get(page, ("", ""))
    co = current_company()
    inputs = _co_inputs(co).get(page)

    # ── Row 1: title (left) + company badge (centre-right) + button (far right) ──
    c_title, c_co, c_btn = st.columns([5, 4, 1])
    with c_title:
        st.markdown(
            f'<div style="padding:6px 0 4px;">'
            f'<div class="page-hdr-title">{title}</div>'
            f'<div class="page-hdr-sub">{subtitle}</div></div>',
            unsafe_allow_html=True)
    with c_co:
        st.markdown(
            f'<div style="padding:8px 0 4px;text-align:right;">'
            f'<div class="page-hdr-co">'
            f'<span class="page-hdr-co-name">{co["name"]}</span>'
            f'<span style="color:#CBD5E1;font-size:.7rem;">|</span>'
            f'<span class="page-hdr-co-sector">{co["sector"]}</span>'
            f'</div></div>',
            unsafe_allow_html=True)
    with c_btn:
        st.markdown('<div style="padding-top:9px;"></div>', unsafe_allow_html=True)
        if st.button("↻", key="reroll_btn", help="Next practice company",
                     use_container_width=True):
            st.session_state.company_idx = (st.session_state.company_idx + 1) % len(COMPANIES)
            st.rerun()

    # ── Row 2: given inputs strip ──
    if inputs is None:
        inputs_list = [("Rev", co["revenue"]), ("EBITDA", co["ebitda"]),
                       ("NI", co["ni"]), ("Price", co["price"]), ("Shares", co["shares"])]
    else:
        inputs_list = inputs
    chips_html = "".join(
        f'<span class="input-chip"><span class="input-chip-lbl">{lbl}</span>{val}</span>'
        for lbl, val in inputs_list
    )
    st.markdown(
        f'<div style="padding:6px 0 14px;border-bottom:1px solid #E5E7EB;margin-bottom:18px;">'
        f'<div class="page-hdr-inputs-lbl">Given Inputs — use these values, do not re-derive</div>'
        f'<div>{chips_html}</div></div>',
        unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="nav-wordmark">FM School</div>', unsafe_allow_html=True)
    nav_btn("Overview", "overview")
    st.markdown('<div class="nav-group-hdr">Foundation</div>', unsafe_allow_html=True)
    nav_btn("①  Three Statement", "3stmt")
    st.markdown('<div class="nav-group-hdr">Valuation</div>', unsafe_allow_html=True)
    nav_btn("②  DCF Model", "dcf")
    nav_btn("③  Comps", "comps")
    nav_btn("④  Precedent Txns", "precedent")
    st.markdown('<div class="nav-group-hdr">Advanced</div>', unsafe_allow_html=True)
    nav_btn("⑤  LBO Model", "lbo")
    nav_btn("⑥  Merger Model", "merger")
    st.markdown('<div class="nav-group-hdr">FP&amp;A</div>', unsafe_allow_html=True)
    nav_btn("⑦  Budget vs Actual", "budget")
    page = st.session_state.current_page
    st.markdown(
        '<div style="margin:20px 16px 0;padding-top:14px;border-top:1px solid #E5E7EB;">'
        '<div style="font-size:.53rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;'
        'color:#9CA3AF;margin-bottom:10px;">Excel Convention</div>'
        '<div style="font-size:.77rem;color:#4B5563;line-height:2.1;">'
        '<span style="display:inline-block;width:9px;height:9px;background:#2563EB;margin-right:8px;vertical-align:middle;"></span>Blue — hardcoded input<br>'
        '<span style="display:inline-block;width:9px;height:9px;background:#374151;margin-right:8px;vertical-align:middle;"></span>Black — formula<br>'
        '<span style="display:inline-block;width:9px;height:9px;background:#16A34A;margin-right:8px;vertical-align:middle;"></span>Green — cross-sheet link<br>'
        '<span style="font-family:monospace;font-size:.7rem;background:#E5E7EB;color:#374151;padding:1px 5px;margin-right:6px;">Ctrl+`</span>toggle formula view'
        '</div>'
        f'<div style="margin-top:14px;padding-top:12px;border-top:1px solid #E5E7EB;">'
        f'<div style="font-size:.53rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;color:#9CA3AF;margin-bottom:6px;">Practice Company</div>'
        f'<div style="font-size:.77rem;color:#4B5563;line-height:1.85;">'
        f'<span style="color:#111827;font-weight:700;display:block;margin-bottom:3px;">{current_company()["name"]}</span>'
        f'{current_company()["sector"]}<br>'
        f'Rev {current_company()["revenue"]} · EBITDA {current_company()["ebitda"]}<br>'
        f'NI {current_company()["ni"]} · {current_company()["price"]} / {current_company()["shares"]} sh'
        '</div></div></div>',
        unsafe_allow_html=True)
    st.markdown('<div style="margin:12px 16px 0;padding-top:12px;border-top:1px solid #E5E7EB;">'
                '<div style="font-size:.53rem;font-weight:700;letter-spacing:.2em;text-transform:uppercase;'
                'color:#9CA3AF;margin-bottom:8px;">Build Options</div></div>',
                unsafe_allow_html=True)
    hints_on = st.session_state.get("hints_on", True)
    if st.button(("💡 Hints: ON  — click to hide" if hints_on else "💡 Hints: OFF — click to show"),
                 key="hint_toggle_sb", use_container_width=True):
        st.session_state.hints_on = not hints_on
        st.rerun()

page_header(page)

# ══════════════════════════════════════════════════════════════════════════════
# OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
if page == "overview":
    st.markdown("# Financial Modeling School")
    st.markdown("Build the seven core Excel models used in investment banking and private equity — from a completely blank spreadsheet. Each module teaches you how professional analysts actually construct these models: the logic, the structure, the conventions, and the formulas.")
    concept("How each module works",
        "<b>Learn tab</b> — Read deeply about how the model is constructed: why each section exists, "
        "how the pieces connect, what bankers actually use each line for, and the thinking behind each formula.<br><br>"
        "<b>Build tab</b> — An interactive spreadsheet with guided sections. Type row labels in col A and "
        "values or formulas in col B. Formulas starting with <b>=</b> are evaluated live — just like Excel. "
        "Guidance on the left of each section directs your thinking without giving answers away.")
    pro_tip("The only way to learn financial modeling is to build models. Reading is essential but not enough — "
            "you need your hands on the keyboard, typing formulas, making mistakes, and fixing them. "
            "Every module here is a build exercise, not a reading exercise.")
    st.markdown("---")
    st.markdown('<div class="sec">Build Order</div>', unsafe_allow_html=True)
    for key, num, name, cat, time, color, desc in [
        ("3stmt","①","Three Statement Model","Foundation","~50 min","#3B82F6",
         "The non-negotiable starting point. Three linked tabs (IS, BS, CFS) in one workbook. Every other model either embeds this or assumes you can read it fluently."),
        ("dcf","②","DCF Model","Valuation","~45 min","#8B5CF6",
         "Project free cash flows, calculate WACC from first principles, build a terminal value, and discount everything to a share price. Add a sensitivity table."),
        ("comps","③","Comparable Company Analysis","Valuation","~30 min","#10B981",
         "Pull public market multiples for 4–6 peers, compute medians, and apply them to NovaTech. The fastest model to build; always paired with DCF."),
        ("precedent","④","Precedent Transactions","Valuation","~25 min","#F59E0B",
         "Same architecture as comps but using M&A deal data. Adds control premium columns. Gives a higher valuation range because acquirers pay to gain control."),
        ("lbo","⑤","LBO Model","Advanced","~55 min","#EF4444",
         "Sources & Uses → debt schedule → five-year hold → exit. MOIC and IRR outputs. The core private equity interview model."),
        ("merger","⑥","Merger Model","Advanced","~40 min","#EC4899",
         "Combine two income statements, issue shares, add synergies, and test whether the deal increases or decreases the acquirer's EPS."),
        ("budget","⑦","Budget vs. Actual","FP&A","~25 min","#14B8A6",
         "Side-by-side variance table with F/U flags, then a volume/price decomposition of the revenue miss. The monthly rhythm of corporate finance."),
    ]:
        st.markdown(
            f'<div class="lp-card" style="border-top-color:{color};">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;">'
            f'<div class="lp-title">{num} {name}</div>'
            f'<div style="font-size:.74rem;color:#9CA3AF;">{cat} · {time}</div></div>'
            f'<div class="lp-desc">{desc}</div></div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# THREE STATEMENT MODEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "3stmt":
    st.markdown("# ① Three Statement Model")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## Why Three Statements?")
        st.markdown("""
Every public company must publish three financial statements. Individually, each tells only part of the story. Together, they give a complete picture of a business's financial reality — and linking them in one Excel workbook is the fundamental skill every financial analyst builds first.

**The Income Statement** answers: was the company profitable over this period? It runs from revenue down to net income across a quarter or year, deducting each category of cost in sequence.

**The Balance Sheet** answers: what does the company own and owe at this exact moment? It is a snapshot — assets on one side, liabilities plus equity on the other — and it must always balance by accounting law.

**The Cash Flow Statement** answers: how did cash actually move? This is the most important statement for understanding business health, because a company can show strong net income while simultaneously running out of cash.

The three-statement *model* links these with live Excel formulas so that changing one assumption — say, raising revenue growth by 2% — automatically ripples through all three statements. That dynamic linkage is what makes it a model rather than three separate spreadsheets.
""")
        st.markdown("## Building the Income Statement")
        st.markdown("""
The Income Statement is built top-down. You always start with **Revenue** — every dollar the company earned from its core operations. Then you deduct costs in a specific order that reveals different profit margins at each level.

**Why this order matters:** Each subtotal tells you something different. Gross Profit isolates the profitability of production. EBITDA isolates operating cash generation. EBIT incorporates asset aging costs. EBT incorporates financing costs. Net Income incorporates taxes. Bankers move up and down this stack constantly — an LBO analyst cares about EBITDA, an equity investor cares about EPS, a credit analyst cares about interest coverage.

**COGS vs. SG&A — a critical distinction.** COGS (Cost of Goods Sold) are variable costs that scale with revenue: raw materials, direct labor, manufacturing overhead. SG&A (Selling, General & Administrative) are largely fixed: office rent, executive salaries, software subscriptions. COGS is modeled as a % of revenue; SG&A is often modeled as a fixed or slowly growing dollar amount.

**D&A — non-cash but real.** Depreciation & Amortization represents the accounting recognition of an asset losing value over time. It reduces reported profit (and thus taxes) but does NOT reduce cash. This is why we add it back in the Cash Flow Statement.
""")
        st.markdown("## Building the Balance Sheet")
        st.markdown("""
The Balance Sheet rests on one identity that must always hold: **Assets = Liabilities + Equity.** If it doesn't balance, you have an error somewhere — and no banker will use a model that doesn't balance.

**Assets** are organized by liquidity — current assets (cash, receivables, inventory) come first because they convert within a year. Then long-term assets (PP&E, intangibles), which represent the company's productive capacity.

**The balance check cell.** Every professional model has a dedicated check cell: =Total Assets − Total L&E. If this cell is not zero, something is broken. Build it immediately and watch it.

**Retained Earnings is the link from the Income Statement.** Each period's Net Income increases Retained Earnings (less any dividends paid). This is one of the key connections between the IS and BS.
""")
        st.markdown("## Building the Cash Flow Statement and Linking the Model")
        st.markdown("""
The CFS is structured in three sections: Operating Activities, Investing Activities, and Financing Activities.

**Operating Activities** starts with Net Income from the IS (cross-sheet link: =IS!B12). Then it makes two types of adjustments:
1. *Add back non-cash charges:* D&A is subtracted on the IS to reduce taxable income, but no cash left the company. So we add it back here.
2. *Adjust for working capital changes:* If Accounts Receivable increased, you recognized revenue but haven't collected the cash yet — that's a negative adjustment. If Accounts Payable increased, you're holding onto cash longer — that's positive.

**Investing Activities** is primarily Capex — spending on PP&E, always entered as a negative number.

**Financing Activities** captures dividends paid, debt raised/repaid, and share buybacks.

**Sheet referencing syntax.** To link between tabs, type =IS!B12 in another tab. This syntax is the backbone of every three-statement model.
""")
        warning("The most common beginner mistake: hard-coding numbers on the CFS instead of linking them from the IS. If you type 52.5 for Net Income instead of =IS!B12, the model breaks as soon as you change any revenue assumption.")
        pro_tip("Build the IS first, then BS, then CFS. Use Ctrl+` to toggle formula view in Excel and verify every non-blue cell is a formula — no exceptions.")

    with build_tab:
        st.markdown("## Build the Three Statement Model")
        st.markdown("Three sections — one per statement. Type labels in col A, values or formulas in col B. "
                    "Formulas starting with **=** evaluate live. The guidance panel on the left of each section gives you hints without giving away the answer.")

        _build_tab_header("3stmt", "Sheets: IS · BS · CFS — Complete (reference) + Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**IS — Revenue:** Hard-code from given inputs. Anchors every formula below it.",
            "**IS — COGS:** Hard-code from inputs. Direct production cost; scales with revenue.",
            "**IS — Gross Profit:** Formula `=B2-B3`. First profitability subtotal.",
            "**IS — SG&A:** Hard-code from inputs. Fixed overhead, separate from production cost.",
            "**IS — EBITDA:** Formula `=B4-B5`. Banker's key operating metric — before D&A, interest, taxes.",
            "**IS — D&A:** Hard-code from inputs. Non-cash charge — CFS will add it back (no cash left the company).",
            "**IS — EBIT:** Formula `=B6-B7`. Operating profit after asset aging.",
            "**IS — Interest Expense:** Hard-code from inputs. Financing cost, not an operating item.",
            "**IS — EBT:** Formula `=B8-B9`. Pre-tax earnings after all costs.",
            "**IS — Tax Rate:** Hard-code as decimal (0.30). Always its own dedicated cell — never embed in a formula.",
            "**IS — Tax Expense:** Formula `=B10*B11`. Reference the rate cell above.",
            "**IS — Net Income:** Formula `=B10-B12`. Note this row number — CFS will cross-link here.",
            "**IS — NI Margin:** Formula `=B13/B2`. Should be single-digit to low-teens percent.",
            "**BS — Assets:** Cash, AR, Inventory are given. Current Assets = `=B2+B3+B4`. PP&E & Intangibles given. Total Assets = `=B5+B6+B7`.",
            "**BS — Liabilities:** AP, ST Debt, LT Debt are given. Total Liabilities = `=B10+B11+B12`.",
            "**BS — Equity:** Common Stock and Retained Earnings are given. Total Equity = `=B15+B16`. Total L&E = `=B13+B17`.",
            "**BS — Balance Check:** Formula `=B8-B18` must equal zero. If not, find the error before using the model.",
            "**CFS — Net Income:** Cross-sheet link from IS tab: `='IS'!B13`. In the Complete sheet we used `='IS — Complete'!B13`.",
            "**CFS — Add: D&A:** Cross-sheet link from IS D&A row. Enter as positive — no cash left the company.",
            "**CFS — Working Capital:** ΔAR growing = negative (cash not yet collected). ΔInventory growing = negative. ΔAP growing = positive (holding cash longer).",
            "**CFS — Cash from Operations:** `=SUM(B3:B7)`. Most important health metric — negative here is a serious warning sign.",
            "**CFS — Capex:** Enter as negative (cash outflow). Cash from Investing = `=B11`.",
            "**CFS — Financing:** Dividends Paid as negative. Cash from Financing = `=B15`.",
            "**CFS — Ending Cash:** `='BS'!B2+B8+B12+B16` — Beginning Cash + Ops + Investing + Financing. Should match BS Cash.",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# DCF MODEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "dcf":
    st.markdown("# ② DCF Model — Discounted Cash Flow")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## What Is a DCF and Why Do Bankers Build It?")
        st.markdown("""
A DCF (Discounted Cash Flow) is an **intrinsic valuation** — it values a company purely on its own fundamentals, independent of what the stock market is currently pricing it at. If the stock is at $45 but your DCF says $60, the company is undervalued. If it says $30, it's overvalued. The DCF is the tool that lets you form that independent view.

The model rests on a single economic principle: a dollar today is worth more than a dollar in the future, because today's dollar can be invested and grow. Therefore, to find what future cash flows are worth *today*, you discount them back at the rate investors require to take on the risk of holding this company's stock and debt. That rate is called WACC.

DCFs are never presented as a single number. Every assumption — growth rate, margins, WACC, terminal growth — is uncertain. Bankers always pair a DCF with a sensitivity table showing how the implied share price changes as assumptions vary.
""")
        st.markdown("## Constructing the Projection Period")
        st.markdown("""
The projection period is typically 5 years. Revenue projections are built from a growth rate assumption. In Excel, you anchor the base-year revenue with an absolute reference ($B$2) and apply compound growth: =$B$2*(1+$B$3)^1 for Year 1, ^2 for Year 2, etc. The dollar signs are essential — they prevent the reference from shifting when you drag the formula right.

**Free Cash Flow** is the number being discounted: FCF = EBITDA × (1 − tax rate) − Capex. This shorthand approximation is common in simplified DCFs.

**Discounting each year's FCF** uses: PV = FCF ÷ (1 + WACC)^n, where n is the year number.
""")
        st.markdown("## Building WACC From First Principles")
        st.markdown("""
**Cost of Equity (CAPM):** Ke = Risk-Free Rate + Beta × Equity Risk Premium. For NovaTech: Ke = 4.5% + 1.1 × 5.5% = 10.55%.

**After-tax Cost of Debt:** Kd × (1 − tax rate). For NovaTech: 6% × (1 − 30%) = 4.2% after-tax.

**Weighting:** WACC = (E/V) × Ke + (D/V) × Kd × (1−T), where V = E + D. Use *market value* weights. For NovaTech: market cap = $900M, debt = $125M, V = $1,025M. WACC ≈ 9.35%.
""")
        st.markdown("## Terminal Value and the Enterprise Value Bridge")
        st.markdown("""
**Gordon Growth Model** (most common): TV = FCF(Y5) × (1 + g) / (WACC − g). The formula assumes FCF grows at g forever. g should be close to long-run nominal GDP growth (2–3%) and MUST be less than WACC or the model produces infinite value.

The TV must then be discounted to present value: PV of TV = TV / (1 + WACC)^5. TV typically represents 60–80% of total DCF value — this is why WACC and g are so sensitive.

**Bridge to share price:**
1. Enterprise Value = Sum of PV(FCFs) + PV(Terminal Value)
2. Equity Value = Enterprise Value − Net Debt
3. Implied Share Price = Equity Value ÷ Shares Outstanding
""")
        warning("Never present a single DCF price as the answer. Always build and present a sensitivity table with WACC on one axis and terminal growth rate on the other.")
        pro_tip("When building the projection block, put Year 1 in col D and Year 5 in col H. Lock assumption references with $ (=$B$3) before dragging — this is the most common formula error in DCF models.")

    with build_tab:
        st.markdown("## Build the DCF Model")
        st.markdown("Three sections: assumptions (col B), 5-year projections (cols B–F per year), then terminal value and the bridge to share price.")

        _build_tab_header("dcf", "Sheets: DCF — Complete (reference) + DCF — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Assumptions block (col B):** Enter Base Revenue, Revenue Growth %, EBITDA Margin %, D&A % Rev, Capex % Rev, Tax Rate, WACC, Terminal g, Net Debt, Shares Out. — all from given inputs. Store each in its own cell.",
            "**Projections — Revenue (row 16, cols C–G):** `=$B$4*(1+$B$5)^n` where n=1 for Yr1, 2 for Yr2, etc. The $ locks prevent the reference from shifting when you drag right.",
            "**Projections — EBITDA (row 17):** `=C16*$B$6` — Revenue × EBITDA Margin. Lock the margin reference with $.",
            "**Projections — Free Cash Flow (row 18):** `=C17*(1-$B$9)-C16*$B$8` — EBITDA×(1-tax) minus Capex. FCF is what gets discounted.",
            "**Projections — Discount Factor (row 19):** `=1/(1+$B$10)^n` — one divided by (1+WACC) raised to the year number.",
            "**Projections — PV of FCF (row 20):** `=C18*C19` — FCF × Discount Factor. Repeat for all 5 years.",
            "**Terminal Value:** `=B23*(1+B11)/(B10-B11)` — Year 5 FCF × (1+g) ÷ (WACC−g). g MUST be less than WACC or the formula produces infinite value.",
            "**PV of Terminal Value:** `=B24/(1+B10)^5` — discount the terminal value back 5 years.",
            "**Enterprise Value:** Sum of PV(FCFs) + PV(TV). PV(FCFs) = `=SUM(C20:G20)`.",
            "**Bridge to Share Price:** EV − Net Debt = Equity Value. Equity Value ÷ Shares Outstanding = Implied Share Price. Compare to the given current stock price.",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# COMPS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "comps":
    st.markdown("# ③ Comparable Company Analysis")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## The Logic of Relative Valuation")
        st.markdown("""
Comps (Comparable Company Analysis) is built on a simple market principle: similar assets should trade at similar prices. If four technology companies with similar growth rates and margins trade at a median of 9.5× EV/EBITDA, a fifth similar company should also trade near 9.5× — unless there's a specific reason it deserves a premium or discount.

This makes comps fast, defensible, and grounded in real market data. Unlike a DCF, it doesn't require you to forecast cash flows — it just asks what the market is currently paying. The challenge is in the peer selection: pick the wrong comparables and your valuation is meaningless.
""")
        st.markdown("## Understanding the Multiples")
        st.markdown("""
**EV/EBITDA** is the most commonly used multiple in investment banking. Enterprise Value (market cap + net debt) divided by EBITDA gives you a measure of how many years of operating earnings you're paying for the whole business. It's capital-structure neutral — both EV and EBITDA are unaffected by whether a company uses debt or equity.

**EV/Revenue** is used when EBITDA is negative or unreliable — early-stage companies, high-growth SaaS, or businesses with deliberately low margins.

**P/E (Price-to-Earnings)** uses market cap (equity value, after debt) divided by net income (also after interest). Two identical businesses with different debt levels will have very different P/Es. Bankers prefer EV-based multiples for cross-company comparisons.

**Why median rather than mean?** One outlier peer trading at 25× in a group that otherwise trades at 9–10× would distort the mean dramatically. The median is robust to outliers.
""")
        st.markdown("## Building the Comps Table in Excel")
        st.markdown("""
One row per comparable company, with metric columns on the right. Col A = company name. Then input columns for raw data (EV, EBITDA, Revenue, Net Income, Market Cap). Then formula columns for the derived multiples.

The summary block uses Excel statistical functions: =MEDIAN(range) and =PERCENTILE(range, 0.25). Apply median EV/EBITDA × NovaTech EBITDA to get implied EV → subtract net debt → divide by shares → implied share price.
""")
        warning("Comps are only as good as your peer selection. Including a company with 40% revenue growth alongside companies growing at 5% will distort the multiple.")

    with build_tab:
        st.markdown("## Build the Comps Table")
        st.markdown("Peer data (all $M): **TechAlpha** EV 1,104 · EBITDA 120 · Rev 480 · NI 65 · MktCap 1,202 | "
                    "**DataCore** EV 810 · EBITDA 100 · Rev 420 · NI 50 · MktCap 810 | "
                    "**CloudSystems** EV 1,260 · EBITDA 120 · Rev 400 · NI 59 · MktCap 1,239 | "
                    "**InfoPro** EV 858 · EBITDA 110 · Rev 510 · NI 54 · MktCap 853")

        _build_tab_header("comps", "Sheets: Comps — Complete (reference) + Comps — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Peer table headers (row 1):** Company · EV · EBITDA · Revenue · Net Income · Mkt Cap · EV/EBITDA · EV/Revenue. Columns for raw data first, then derived multiples.",
            "**Peer data rows 2–5:** Enter all raw data (EV, EBITDA, Revenue, NI, Mkt Cap) as blue given inputs. EV/EBITDA = `=B2/C2`. EV/Revenue = `=B2/D2`. Copy formulas down.",
            "**Statistics block:** 25th Percentile = `=PERCENTILE(G2:G5,0.25)`. Median = `=MEDIAN(G2:G5)`. 75th Pctile = `=PERCENTILE(G2:G5,0.75)`. Repeat for EV/Revenue column.",
            "**Implied Valuation:** Enter subject company EBITDA (given). Reference Median EV/EBITDA from stats block (do not retype). Implied EV = EBITDA × Median multiple.",
            "**Bridge to share price:** Implied EV − Net Debt = Equity Value. Equity Value ÷ Shares Out = Implied Share Price. Compare to current price — is the stock cheap or expensive vs. peers?",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# PRECEDENT TRANSACTIONS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "precedent":
    st.markdown("# ④ Precedent Transaction Analysis")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## Why Precedent Transactions Give Higher Valuations Than Comps")
        st.markdown("""
Precedent transaction analysis uses the same multiple framework as comps — but instead of trading multiples (what the market pays for a minority stake today), it uses acquisition multiples (what acquirers paid to buy entire companies in completed M&A deals). These are almost always higher for one reason: **the control premium**.

When an acquirer buys a company, they're buying control of the entire enterprise — the ability to install management, cut costs, redirect capital. Shareholders won't sell without receiving a premium above the current stock price. Typical control premiums run 20–40%.
""")
        st.markdown("## Construction Differences from Comps")
        st.markdown("""
**Unaffected stock price.** This is the target's stock price 30 days before any deal rumors. You use it because by announcement, the stock has already jumped. Using the post-announcement price would understate the control premium.

**Control premium column.** Calculated as: (Deal EV − Unaffected Market Cap) ÷ Unaffected Market Cap. Apply the average premium to NovaTech's current price to estimate a minimum acquisition price.

**LTM EBITDA at time of deal, not current.** When looking at the multiple an acquirer paid, use the EBITDA that existed at deal time — not today's figure.
""")
        warning("Precedents go stale quickly. A 2015 deal multiple reflects 2015 market conditions. Weight recent transactions more heavily and exclude deals from unusually hot or distressed markets.")

    with build_tab:
        st.markdown("## Build the Precedent Transactions Table")
        st.markdown("Deal data (all $M): **AcquireCo/TechAlpha '22** Deal EV 1,320 · LTM EBITDA 120 · Pre-deal MktCap 975 | "
                    "**MegaCorp/DataCore '21** Deal EV 980 · LTM EBITDA 100 · Pre-deal MktCap 809 | "
                    "**GlobalTech/InfoPro '23** Deal EV 1,045 · LTM EBITDA 110 · Pre-deal MktCap 856")

        _build_tab_header("precedent", "Sheets: Precedent — Complete (reference) + Precedent — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Deal table headers (row 1):** Deal · Year · Deal EV · LTM EBITDA · EV/EBITDA · Pre-deal Mkt Cap. The control premium column captures the takeover premium paid.",
            "**Deal data rows 2–4:** Enter Deal EV, LTM EBITDA, Pre-deal Mkt Cap as given. EV/EBITDA = Deal EV ÷ LTM EBITDA. These multiples should be 10–35% higher than trading comps — that's the control premium.",
            "**Statistics:** Median EV/EBITDA = `=MEDIAN(E2:E4)`. Compare to your comps median — precedents should be higher.",
            "**Implied Valuation:** Subject EBITDA × Median Deal Multiple = Implied EV. Then EV − Net Debt ÷ Shares = Implied Share Price.",
            "**Key distinction:** These are ACQUISITION multiples (control premium included), not trading multiples. A company's acquisition floor price is always above its trading price.",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# LBO
# ══════════════════════════════════════════════════════════════════════════════
elif page == "lbo":
    st.markdown("# ⑤ LBO Model — Leveraged Buyout")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## What Is an LBO and Why Does Leverage Drive Returns?")
        st.markdown("""
An LBO (Leveraged Buyout) is an acquisition financed primarily with borrowed money — typically 60% debt, 40% equity. A private equity firm provides the equity and uses the acquisition target's own cash flows to service and pay down the debt over a 3–7 year hold period. At exit (a sale or IPO), the firm collects the remaining equity value.

Why leverage? Because it amplifies returns on equity invested. If you buy a $1,080M company with $432M equity and $648M debt, and the business generates enough cash flow to pay down debt over five years, the equity at exit is worth far more than the $432M invested — even if the exit multiple stays the same.
""")
        st.markdown("## The Three Sources of PE Return")
        st.markdown("""
**1. EBITDA Growth.** The company becomes more profitable — through revenue growth, cost cuts, or add-on acquisitions.

**2. Deleveraging.** Every dollar of cash flow used to pay down debt increases equity value by a dollar (Equity = EV − Debt). A company that pays down $200M of debt over five years creates $200M of additional equity value even if EV stays flat.

**3. Multiple Expansion.** Buying at 9× and selling at 11× captures the difference. This is the least reliable source — it depends on market conditions at exit — so conservative models assume entry multiple = exit multiple.
""")
        st.markdown("## Building the Model: Three Sections")
        st.markdown("""
**Section 1 — Sources & Uses.** Purchase price = entry multiple × LTM EBITDA. Structure: 60% debt, 40% PE equity. Sources must equal Uses.

**Section 2 — Debt Schedule.** Tracks the loan balance year by year. Each year: Beginning Debt → subtract cash available for paydown → Ending Debt → feeds next year's Beginning Debt. This rolling structure is the core of the model.

**Section 3 — Exit & Returns.** Exit EV = Year 5 EBITDA × exit multiple. Exit Equity = Exit EV − remaining debt. MOIC = Exit Equity ÷ Entry Equity. IRR uses Excel's =IRR() function on a cash flow row: −entry equity at year 0, exit equity at year 5, zeros in between.
""")
        warning("The exit multiple assumption is the single biggest variable in an LBO. Always assume exit = entry in your base case. Multiple expansion is a bonus, not a plan.")
        pro_tip("Build the debt schedule as a rolling structure: Ending Debt in Yr1 = Beginning Debt − paydown, and Beginning Debt in Yr2 = Yr1 Ending Debt. This is more robust than a one-line formula.")

    with build_tab:
        st.markdown("## Build the LBO Model")

        _build_tab_header("lbo", "Sheets: LBO — Complete (reference) + LBO — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Sources & Uses:** LTM EBITDA × Entry Multiple = Entry EV. Debt = EV × 60%. PE Equity = EV × 40%. Sources Check = Debt + Equity (must equal Entry EV).",
            "**Debt Schedule — Beginning Debt (row 12):** Yr1 Beginning Debt = Entry Debt from S&U (`=B6`). Yr2+ Beginning Debt = prior year's Ending Debt. This rolling structure tracks deleveraging year by year.",
            "**Debt Schedule — Interest (row 13):** Beginning Debt × Interest Rate (7%). Use the rate from inputs.",
            "**Debt Schedule — EBITDA (row 14):** Entry EBITDA × (1+growth)^n — compound growth each year.",
            "**Debt Schedule — FCF for Paydown (row 15):** EBITDA × (1-Tax) − Capex. This is the cash available to pay down debt.",
            "**Debt Schedule — Ending Debt (row 16):** Beginning Debt − FCF for Paydown. Watch this fall each year — that's deleveraging.",
            "**Exit & Returns:** Exit EV = Exit Multiple × Year 5 EBITDA. Exit Equity = Exit EV − Year 5 Remaining Debt. MOIC = Exit Equity ÷ Entry Equity.",
            "**IRR:** Build a cash flow row: [−Entry Equity, 0, 0, 0, 0, Exit Equity]. Apply `=IRR(row)`. Above 20% is the institutional PE benchmark.",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# MERGER MODEL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "merger":
    st.markdown("# ⑥ Merger Model — M&A Accretion / Dilution")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## What Is Accretion and Why Do Boards Care?")
        st.markdown("""
A merger model answers one question: if Company A acquires Company B, does the combined company's Earnings Per Share (EPS) go up or down relative to Company A's standalone EPS?

**Accretive** means the deal increases EPS. **Dilutive** means the deal decreases EPS. Boards and public investors focus intensely on this metric because EPS drives stock price. A dilutive deal — especially one without a compelling strategic rationale — invites shareholder criticism and activist pressure.
""")
        st.markdown("## Why All-Stock Deals Are More Complex")
        st.markdown("""
In a cash deal, the acquirer simply pays cash and the target's shareholders disappear. No new shares are issued.

In an all-stock deal, the acquirer issues new shares to pay for the target. This increases the share count, which dilutes EPS unless the earnings contribution from the acquired business more than compensates. Fewer new shares = less dilution. A higher acquirer stock price means fewer new shares need to be issued for the same deal value.
""")
        st.markdown("## Building the Pro Forma Income Statement")
        st.markdown("""
**Synergies (positive).** Cost synergies from eliminating duplicate functions. Enter as a positive number — they increase combined earnings.

**Amortization of Acquired Intangibles (negative).** When you acquire a company, accounting rules require you to step up intangible assets to fair value. These are then amortized — a recurring non-cash expense that hits reported EPS. Enter as negative.

**The EPS test.** Pro Forma EPS = Pro Forma Net Income ÷ Pro Forma Shares. Compare to standalone EPS = Acquirer NI ÷ Acquirer Shares. % change = (PF EPS − Standalone EPS) ÷ Standalone EPS.
""")
        warning("Synergies are almost always overstated by management. Use conservative estimates in base case and show a sensitivity for break-even synergies — the minimum needed for the deal to be EPS-neutral.")

    with build_tab:
        st.markdown("## Build the Merger Model")

        _build_tab_header("merger", "Sheets: Merger — Complete (reference) + Merger — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Deal Assumptions:** Target Market Price × (1 + Premium) = Offer Price/Share. Total Deal Value = Offer Price × Target Shares. New Shares Issued = Total Deal Value ÷ Acquirer Price.",
            "**Pro Forma Income Statement:** Acquirer NI + Target NI + After-tax Synergies − Intangibles Amortization = Pro Forma NI. Synergies are positive; amortization is negative.",
            "**EPS Test — Standalone EPS:** Acquirer NI ÷ Acquirer Shares. This is the benchmark you're testing against.",
            "**EPS Test — Pro Forma EPS:** Pro Forma NI ÷ (Acquirer Shares + New Shares Issued). Compare to Standalone EPS.",
            "**Accretion/Dilution:** `=(PF EPS − Standalone EPS) / Standalone EPS`. Positive = accretive (good). Negative = dilutive (management must explain why the deal is strategically worth the EPS hit).",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# BUDGET VS. ACTUAL
# ══════════════════════════════════════════════════════════════════════════════
elif page == "budget":
    st.markdown("# ⑦ Budget vs. Actual Model")
    learn_tab, build_tab = st.tabs(["Learn the Model", "Build It in Excel"])

    with learn_tab:
        st.markdown("## What BvA Is and Why It Matters")
        st.markdown("""
Budget vs. Actual (BvA) is the core rhythm of corporate finance. Every month, the FP&A team publishes a report comparing what was planned (the budget) against what actually happened (the actuals). The result is a variance report.

The skill is in the interpretation. A $7M revenue miss sounds bad. But is it because the sales team sold fewer units (a volume problem)? Or because they sold the right number at lower prices (a pricing problem)? These require completely different responses. The BvA model's job is to separate them.
""")
        st.markdown("## Building the Variance Table")
        st.markdown("""
The table architecture: one row per P&L line item (Revenue, COGS, Gross Profit, SG&A, EBITDA), with columns for Budget, Actual, Variance $, Variance %, and a Favorable/Unfavorable flag.

**Variance $ = Actual − Budget.** For revenue, positive is good. For costs, positive means you overspent.

**Gross Profit and EBITDA are derived, not input.** Budget GP = =B2-B3. Actual GP = =C2-C3. If a line is a formula in budget, it must also be a formula in actual.

**F/U flag:** for revenue =IF(Variance>=0,"F","U"). For cost rows, the logic reverses.
""")
        st.markdown("## Volume / Price Decomposition")
        st.markdown("""
**Volume Effect = (Actual Units − Budget Units) × Budget Price.** Isolates the impact of selling fewer units, holding price constant.

**Price Effect = Actual Units × (Actual Price − Budget Price).** Isolates pricing impact at actual volume.

Volume Effect + Price Effect = Total Revenue Variance. This is your built-in check.
""")
        warning("A revenue miss almost always creates a disproportionately larger EBITDA miss. If revenue misses by 5%, EBITDA might miss by 25% — because many costs are fixed.")
        pro_tip("Apply conditional formatting to the Variance $ column: red for negative values, green for positive. This makes the table instantly scannable in a CFO presentation.")

    with build_tab:
        st.markdown("## Build the Budget vs. Actual Model")
        st.markdown("Q1 data: Budget revenue $120M, actual $113M · Budget COGS $72M, actual $70.3M · Budget SG&A $20M, actual $21.5M")

        _build_tab_header("budget", "Sheets: BvA — Complete (reference) + BvA — Practice (write your formulas)")
        st.markdown("### Build Guidance")
        for _i, _h in enumerate([
            "**Main table headers:** Line Item · Budget · Actual · Variance $ · Variance % · F/U. One row per P&L line.",
            "**Revenue row:** Budget and Actual are given (blue). Variance $ = `=Actual-Budget`. For revenue, positive = Favorable: `=IF(D2>=0,\"F\",\"U\")`.",
            "**COGS row:** Given inputs. F/U logic FLIPS for costs: `=IF(D3<=0,\"F\",\"U\")` — lower cost = Favorable.",
            "**Gross Profit row:** NEVER hard-code. Budget = `=B2-B3`. Actual = `=C2-C3`. Then apply Variance $ and F/U formulas.",
            "**SG&A:** Given inputs. Same F/U flip as COGS. EBITDA: formula derived row, same as Gross Profit.",
            "**EBITDA Margin:** `=EBITDA/Revenue` for each column. Notice: a ~6% revenue miss causes a much larger EBITDA % miss — fixed costs amplify the variance.",
            "**Volume/Price Decomp:** Volume Effect = `=(Actual Units − Budget Units) × Budget Price`. Price Effect = `=Actual Units × (Actual Price − Budget Price)`. Sum must equal Revenue Variance $.",
        ], 1):
            st.markdown(f'<div class="guide-line"><span class="hint-num">{_i}.</span>{_h}</div>', unsafe_allow_html=True)
